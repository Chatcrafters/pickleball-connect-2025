from urllib.parse import quote
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify
from models import db, PCLTournament, PCLTeam, PCLRegistration, Player, SHIRT_SIZES, COUNTRY_FLAGS, get_whatsapp_sponsor_block, PCLMatch, PCLLineup, PCLMatchResult, PoolPlayer
from datetime import datetime, date
from werkzeug.utils import secure_filename
from utils.supabase_storage import upload_photo_to_supabase, get_photo_url
from utils.whatsapp import send_whatsapp_message, send_captain_invitation_template
import base64
import os
import csv
import io
import json
from urllib.parse import quote

pcl = Blueprint('pcl', __name__)

# Configuration for file uploads
UPLOAD_FOLDER = 'static/uploads/pcl'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ============================================================================
# TRANSLATIONS (Fixed UTF-8 + Quick Add Extensions)
# ============================================================================

TRANSLATIONS = {
    'EN': {
        'page_title': 'PCL Player Registration',
        'team': 'Team',
        'personal_info': 'Personal Information',
        'first_name': 'First Name',
        'last_name': 'Last Name',
        'email': 'Email',
        'phone': 'Phone',
        'phone_help': 'With country code (+49, +34, etc.)',
        'gender': 'Gender',
        'male': 'Male',
        'female': 'Female',
        'birth_year': 'Birth Year',
        'role': 'Role',
        'player': 'Player',
        'captain': 'Captain',
        'shirt_info': 'Shirt Information',
        'shirt_size': 'Shirt Size',
        'shirt_second': 'Second Shirt (optional)',
        'shirt_third': 'Third Shirt (optional)',
        'shirt_total_price': 'Total',
        'shirt_payment_info': 'Each shirt costs 15 EUR. Shirts are ordered and printed by us according to official PCL specifications. Payment is settled through your team captain.',
        'shirt_product_preview_helper': "This is the shirt you'll receive. Choose your size below.",
        'match_status_pending': 'Pending',
        'match_status_lineups_locked': 'Lineups locked',
        'match_status_in_progress': 'In progress',
        'match_status_completed': 'Completed',
        'match_create_new': 'New Match',
        'match_team_home': 'Home',
        'match_team_away': 'Away',
        'match_court': 'Court',
        'match_lineup_deadline': 'Lineup deadline',
        'match_lineups_submitted': 'Lineups submitted',
        'match_delete_confirm': 'Really delete match?',
        'lineup_submit_title': 'Submit Lineup',
        'lineup_edit_title': 'Edit Lineup',
        'lineup_section_wd': "Women's Doubles",
        'lineup_section_md': "Men's Doubles",
        'lineup_section_mx1': 'Mixed 1',
        'lineup_section_mx2': 'Mixed 2',
        'lineup_section_hb': 'Heartbreaker Order',
        'lineup_select_male': 'Select male',
        'lineup_select_female': 'Select female',
        'lineup_position': 'Position',
        'lineup_status_selected': 'selected',
        'lineup_submit_button': 'Submit Lineup',
        'lineup_update_button': 'Update Lineup',
        'lineup_already_submitted': 'Lineup submitted',
        'lineup_opponent_submitted': 'Opponent submitted',
        'lineup_opponent_pending': 'Opponent not submitted yet',
        'lineup_deadline_in': 'Deadline in',
        'lineup_deadline_passed': 'Deadline passed',
        'lineup_reveal_title': 'Lineups Revealed',
        'lineup_heartbreaker_help': '2 men + 2 women in playing order. Used if match reaches 2:2.',
        'lineup_hb_men_count': 'men',
        'lineup_hb_women_count': 'women',
        'lineup_hb_gender_balance_error': 'Heartbreaker needs exactly 2 men and 2 women',
        'lineup_mixed2_different_warning': 'Mixed 2 cannot use any player from Mixed 1',
        'upcoming_matches_heading': 'Upcoming Matches',
        'past_matches_heading': 'Past Matches',
        'admin_override_mode': 'Admin Override Mode',
        'admin_edit_lineup': 'View/Edit Lineup',
        'admin_submit_for_team': 'Submit Lineup for',
        'score_entry_title': 'Enter Score',
        'score_match_in_progress': 'Match in progress',
        'score_match_completed': 'Match completed',
        'score_heartbreaker_required': 'Heartbreaker required',
        'score_heartbreaker_section': 'Heartbreaker - 1 set to 21',
        'score_wins': 'wins',
        'score_home_team_score': 'Home score',
        'score_away_team_score': 'Away score',
        'score_save': 'Save',
        'score_edit': 'Edit',
        'score_refresh_standing': 'Refresh standing',
        'score_winner': 'Winner',
        'standings_heading': 'Standings',
        'standings_position': 'Pos',
        'standings_team': 'Team',
        'standings_played': 'P',
        'standings_wins': 'W',
        'standings_losses': 'L',
        'standings_points': 'Pts',
        'standings_diff': 'Diff',
        'standings_no_matches_yet': 'No completed matches yet',
        'your_team_indicator': 'Your team',
        'captain_pool_section': 'Pool players in',
        'profile': 'Profile',
        'photo': 'Profile Photo',
        'photo_help': 'Required. JPG, PNG, max 5MB. Square format recommended.',
        'upload_photo': 'Upload Photo',
        'bio': 'Short Bio',
        'bio_placeholder': 'Tell us about yourself, your pickleball journey...',
        'bio_help': 'Tell us about yourself (50-500 characters)',
        'social_media': 'Social Media',
        'optional_info': 'Optional Information',
        'video_url': 'Video URL (Highlight Reel)',
        'dupr_rating': 'DUPR Rating',
        'language': 'Preferred Language',
        'privacy_accept': 'I accept the data processing for PCL registration',
        'submit': 'Register',
        'update': 'Update Registration',
        'required': 'Required',
        'optional': 'Optional',
        'select': 'Select',
        'success_title': 'Registration Complete!',
        'success_message': 'Thank you for registering for PCL.',
        'registration_whatsapp_sent': 'Registration successful! Check WhatsApp for your personal profile link.',
        'captain_is_playing_label': 'Captain also plays in the team',
        'missing_fields': 'Please complete all required fields',
        'captain_dashboard': 'Captain Dashboard',
        'team_status': 'Team Status',
        'registration_link': 'Registration Link for Players',
        'copy_link': 'Copy Link',
        'link_copied': 'Link copied!',
        'players_registered': 'Players',
        'men': 'Men',
        'women': 'Women',
        'complete': 'Complete',
        'incomplete': 'Incomplete',
        'photo_missing': 'Photo missing',
        'deadline': 'Deadline',
        'days_left': 'days left',
        'send_reminder': 'Send Reminder',
        'export_data': 'Export Team Data',
        'missing': 'Missing',
        'no_players': 'No players yet',
        'requirements': 'Requirements',
        'team_size': 'Team Size',
        'required_per_player': 'Required per Player',
        'captain_link_warning': 'This link is only for you as captain. Do not share!',
        # Quick Add specific
        'quick_add_title': 'Quick Add Player',
        'quick_add_player': 'Add Player',
        'quick_add_info': 'The player will complete their profile via WhatsApp link',
        'add_player': 'Add Player',
        'back_to_dashboard': 'Back to Dashboard',
        'send_link_now': 'Send WhatsApp link immediately',
        'send_link_help': 'Player receives a link to complete their profile',
        'phone_invalid': 'Please enter a valid phone number with country code',
        'quick_actions': 'Quick Actions',
        'send_all_links': 'Send All Links',
        'send_all_confirm': 'Send links to all incomplete players?',
        'player_added': 'Player added successfully!',
        'player_added_whatsapp': 'Player added and WhatsApp sent!',
        'whatsapp_sent': 'WhatsApp sent!',
        'whatsapp_failed': 'WhatsApp could not be sent',
        # Complete Profile specific
        'complete_profile_title': 'Complete Your Profile',
        'hello': 'Hello',
        'profile_progress': 'Profile Progress',
        'save_profile': 'Save Profile',
        'profile_saved': 'Profile saved successfully!',
        'profile_updated': 'Profile updated!',
        'edit_profile_title': 'Edit Profile',
        'current_photo': 'Current photo',
        'leave_empty_to_keep': 'Leave empty to keep',
        'data_protection': 'Your data is securely stored',
        'contact': 'Contact',
        'lock_closed_banner': 'Registration closed for this team. Contact Sergio to request changes.',
        'lock_open_until_banner': 'Registration open until {date}',
        'lock_open_manual_banner': 'Registration open (manually unlocked)',
        'lock_disabled_button': 'Locked',
        'lock_admin_section': 'Registration Lock',
        'lock_admin_toggle': 'Registration Open (manual)',
        'lock_admin_deadline': 'Registration Open Until',
        'lock_admin_helper': 'When checked OR until-date is in future, captain can add/edit/delete players.',
    },
    'DE': {
        'page_title': 'PCL Spieler-Registrierung',
        'team': 'Team',
        'personal_info': 'Persoenliche Informationen',
        'first_name': 'Vorname',
        'last_name': 'Nachname',
        'email': 'E-Mail',
        'phone': 'Telefon',
        'phone_help': 'Mit Laendervorwahl (+49, +34, etc.)',
        'gender': 'Geschlecht',
        'male': 'Maennlich',
        'female': 'Weiblich',
        'birth_year': 'Geburtsjahr',
        'role': 'Rolle',
        'player': 'Spieler',
        'captain': 'Kapitaen',
        'shirt_info': 'Shirt-Informationen',
        'shirt_size': 'Shirt-GrÃ¶ÃŸe',
        'shirt_second': 'Zweites Shirt (optional)',
        'shirt_third': 'Drittes Shirt (optional)',
        'shirt_total_price': 'Gesamt',
        'shirt_payment_info': 'Jedes Shirt kostet 15 EUR. Die Shirts werden von uns gemÃ¤ÃŸ den offiziellen PCL-Vorgaben bestellt und bedruckt. Die Abrechnung erfolgt Ã¼ber den TeamkapitÃ¤n.',
        'shirt_product_preview_helper': "Das ist das Shirt, das du erhältst. Wähle unten deine Größe.",
        'match_status_pending': 'Ausstehend',
        'match_status_lineups_locked': 'Aufstellungen gesperrt',
        'match_status_in_progress': 'Im Spiel',
        'match_status_completed': 'Beendet',
        'match_create_new': 'Neue Begegnung',
        'match_team_home': 'Heim',
        'match_team_away': 'Auswaerts',
        'match_court': 'Court',
        'match_lineup_deadline': 'Lineup-Deadline',
        'match_lineups_submitted': 'Aufstellungen eingereicht',
        'match_delete_confirm': 'Begegnung wirklich loeschen?',
        'lineup_submit_title': 'Aufstellung einreichen',
        'lineup_edit_title': 'Aufstellung bearbeiten',
        'lineup_section_wd': 'Damen-Doppel',
        'lineup_section_md': 'Herren-Doppel',
        'lineup_section_mx1': 'Mixed 1',
        'lineup_section_mx2': 'Mixed 2',
        'lineup_section_hb': 'Heartbreaker-Reihenfolge',
        'lineup_select_male': 'Mann waehlen',
        'lineup_select_female': 'Frau waehlen',
        'lineup_position': 'Position',
        'lineup_status_selected': 'ausgewaehlt',
        'lineup_submit_button': 'Aufstellung einreichen',
        'lineup_update_button': 'Aufstellung aktualisieren',
        'lineup_already_submitted': 'Aufstellung eingereicht',
        'lineup_opponent_submitted': 'Gegner hat eingereicht',
        'lineup_opponent_pending': 'Gegner noch nicht eingereicht',
        'lineup_deadline_in': 'Deadline in',
        'lineup_deadline_passed': 'Deadline abgelaufen',
        'lineup_reveal_title': 'Aufstellungen sichtbar',
        'lineup_heartbreaker_help': '2 Maenner + 2 Frauen in Reihenfolge. Wird bei 2:2 verwendet.',
        'lineup_hb_men_count': 'Maenner',
        'lineup_hb_women_count': 'Frauen',
        'lineup_hb_gender_balance_error': 'Heartbreaker braucht genau 2 Maenner und 2 Frauen',
        'lineup_mixed2_different_warning': 'Mixed 2 darf keinen Spieler aus Mixed 1 enthalten',
        'upcoming_matches_heading': 'Kommende Begegnungen',
        'past_matches_heading': 'Vergangene Begegnungen',
        'admin_override_mode': 'Admin-Override-Modus',
        'admin_edit_lineup': 'Aufstellung ansehen/bearbeiten',
        'admin_submit_for_team': 'Aufstellung einreichen fuer',
        'score_entry_title': 'Ergebnis eingeben',
        'score_match_in_progress': 'Spiel laeuft',
        'score_match_completed': 'Spiel beendet',
        'score_heartbreaker_required': 'Heartbreaker erforderlich',
        'score_heartbreaker_section': 'Heartbreaker - 1 Satz bis 21',
        'score_wins': 'gewinnt',
        'score_home_team_score': 'Heim-Punktzahl',
        'score_away_team_score': 'Auswaerts-Punktzahl',
        'score_save': 'Speichern',
        'score_edit': 'Bearbeiten',
        'score_refresh_standing': 'Spielstand aktualisieren',
        'score_winner': 'Sieger',
        'standings_heading': 'Tabelle',
        'standings_position': 'Pos',
        'standings_team': 'Team',
        'standings_played': 'Sp.',
        'standings_wins': 'S',
        'standings_losses': 'N',
        'standings_points': 'Pkt',
        'standings_diff': 'Diff',
        'standings_no_matches_yet': 'Noch keine beendeten Spiele',
        'your_team_indicator': 'Dein Team',
        'captain_pool_section': 'Pool-Spieler in',
        'profile': 'Profil',
        'photo': 'Profilbild',
        'photo_help': 'Pflichtfeld. JPG, PNG, max 5MB. Quadratisches Format empfohlen.',
        'upload_photo': 'Foto hochladen',
        'bio': 'Kurze Bio',
        'bio_placeholder': 'Erzaehl uns von dir und deiner Pickleball-Reise...',
        'bio_help': 'Erzaehl uns von dir (50-500 Zeichen)',
        'social_media': 'Social Media',
        'optional_info': 'Optionale Informationen',
        'video_url': 'Video-URL (Highlight-Video)',
        'dupr_rating': 'DUPR Rating',
        'language': 'Bevorzugte Sprache',
        'privacy_accept': 'Ich stimme der Datenverarbeitung fuer die PCL-Registrierung zu',
        'submit': 'Registrieren',
        'update': 'Registrierung aktualisieren',
        'required': 'Pflicht',
        'optional': 'Optional',
        'select': 'Auswaehlen',
        'success_title': 'Registrierung erfolgreich!',
        'success_message': 'Danke fuer deine Registrierung zur PCL.',
        'registration_whatsapp_sent': 'Registrierung erfolgreich! Schau auf WhatsApp fuer deinen persoenlichen Profil-Link.',
        'captain_is_playing_label': 'Captain spielt auch mit',
        'missing_fields': 'Bitte fuelle alle Pflichtfelder aus',
        'captain_dashboard': 'Kapitaen Dashboard',
        'team_status': 'Team-Status',
        'registration_link': 'Registrierungslink fuer Spieler',
        'copy_link': 'Link kopieren',
        'link_copied': 'Link kopiert!',
        'players_registered': 'Spieler',
        'men': 'Maenner',
        'women': 'Frauen',
        'complete': 'Vollstaendig',
        'incomplete': 'Unvollstaendig',
        'photo_missing': 'Foto fehlt',
        'deadline': 'Deadline',
        'days_left': 'Tage uebrig',
        'send_reminder': 'Erinnerung senden',
        'export_data': 'Team-Daten exportieren',
        'missing': 'Fehlt',
        'no_players': 'Noch keine Spieler',
        'requirements': 'Anforderungen',
        'team_size': 'Teamgroesse',
        'required_per_player': 'Pro Spieler erforderlich',
        'captain_link_warning': 'Dieser Link ist nur fuer dich als Kapitaen. Nicht teilen!',
        # Quick Add specific
        'quick_add_title': 'Spieler schnell hinzufuegen',
        'quick_add_player': 'Spieler hinzufuegen',
        'quick_add_info': 'Der Spieler vervollstaendigt sein Profil ueber den WhatsApp-Link',
        'add_player': 'Spieler hinzufuegen',
        'back_to_dashboard': 'Zurueck zum Dashboard',
        'send_link_now': 'WhatsApp-Link sofort senden',
        'send_link_help': 'Spieler erhaelt Link zur Profil-Vervollstaendigung',
        'phone_invalid': 'Bitte gueltige Telefonnummer mit Laendervorwahl eingeben',
        'quick_actions': 'Schnellaktionen',
        'send_all_links': 'Alle Links senden',
        'send_all_confirm': 'Links an alle unvollstaendigen Spieler senden?',
        'player_added': 'Spieler erfolgreich hinzugefuegt!',
        'player_added_whatsapp': 'Spieler hinzugefuegt und WhatsApp gesendet!',
        'whatsapp_sent': 'WhatsApp gesendet!',
        'whatsapp_failed': 'WhatsApp konnte nicht gesendet werden',
        # Complete Profile specific
        'complete_profile_title': 'Profil vervollstaendigen',
        'hello': 'Hallo',
        'profile_progress': 'Profil-Fortschritt',
        'save_profile': 'Profil speichern',
        'profile_saved': 'Profil erfolgreich gespeichert!',
        'profile_updated': 'Profil aktualisiert!',
        'edit_profile_title': 'Profil bearbeiten',
        'current_photo': 'Aktuelles Foto',
        'leave_empty_to_keep': 'Leer lassen um beizubehalten',
        'data_protection': 'Deine Daten werden sicher gespeichert',
        'contact': 'Kontakt',
        'lock_closed_banner': 'Registrierung fuer dieses Team geschlossen. Kontaktiere Sergio fuer Aenderungen.',
        'lock_open_until_banner': 'Registrierung offen bis {date}',
        'lock_open_manual_banner': 'Registrierung offen (manuell freigeschaltet)',
        'lock_disabled_button': 'Gesperrt',
        'lock_admin_section': 'Registrierungssperre',
        'lock_admin_toggle': 'Registrierung offen (manuell)',
        'lock_admin_deadline': 'Registrierung offen bis',
        'lock_admin_helper': 'Wenn aktiviert ODER Datum in der Zukunft, kann der Kapitaen Spieler hinzufuegen/bearbeiten/loeschen.',
    },
    'ES': {
        'page_title': 'Registro de Jugadores PCL',
        'team': 'Equipo',
        'personal_info': 'InformaciÃƒÂ³n Personal',
        'first_name': 'Nombre',
        'last_name': 'Apellido',
        'email': 'Correo electrÃƒÂ³nico',
        'phone': 'Telefono',
        'phone_help': 'Con cÃƒÂ³digo de pais (+34, +49, etc.)',
        'gender': 'Genero',
        'male': 'Masculino',
        'female': 'Femenino',
        'birth_year': 'AÃƒÂ±o de nacimiento',
        'role': 'Rol',
        'player': 'Jugador',
        'captain': 'CapitÃƒÂ¡n',
        'shirt_info': 'InformaciÃ³n de la Camiseta',
        'shirt_size': 'Talla de camiseta',
        'shirt_second': 'Segunda camiseta (opcional)',
        'shirt_third': 'Tercera camiseta (opcional)',
        'shirt_total_price': 'Total',
        'shirt_payment_info': 'Cada camiseta cuesta 15 EUR. Las camisetas son pedidas e impresas por nosotros segÃºn las especificaciones oficiales de la PCL. El pago se gestiona a travÃ©s del capitÃ¡n del equipo.',
        'shirt_product_preview_helper': "Esta es la camiseta que recibirás. Elige tu talla abajo.",
        'match_status_pending': 'Pendiente',
        'match_status_lineups_locked': 'Alineaciones bloqueadas',
        'match_status_in_progress': 'En curso',
        'match_status_completed': 'Finalizado',
        'match_create_new': 'Nuevo partido',
        'match_team_home': 'Local',
        'match_team_away': 'Visitante',
        'match_court': 'Court',
        'match_lineup_deadline': 'Fecha limite alineacion',
        'match_lineups_submitted': 'Alineaciones enviadas',
        'match_delete_confirm': 'Borrar partido?',
        'lineup_submit_title': 'Enviar alineacion',
        'lineup_edit_title': 'Editar alineacion',
        'lineup_section_wd': 'Dobles femenino',
        'lineup_section_md': 'Dobles masculino',
        'lineup_section_mx1': 'Mixto 1',
        'lineup_section_mx2': 'Mixto 2',
        'lineup_section_hb': 'Orden Heartbreaker',
        'lineup_select_male': 'Elegir hombre',
        'lineup_select_female': 'Elegir mujer',
        'lineup_position': 'Posicion',
        'lineup_status_selected': 'seleccionado',
        'lineup_submit_button': 'Enviar alineacion',
        'lineup_update_button': 'Actualizar alineacion',
        'lineup_already_submitted': 'Alineacion enviada',
        'lineup_opponent_submitted': 'Rival ha enviado',
        'lineup_opponent_pending': 'Rival aun no ha enviado',
        'lineup_deadline_in': 'Fecha limite en',
        'lineup_deadline_passed': 'Fecha limite pasada',
        'lineup_reveal_title': 'Alineaciones reveladas',
        'lineup_heartbreaker_help': '2 hombres + 2 mujeres en orden. Se usa si el partido llega a 2:2.',
        'lineup_hb_men_count': 'hombres',
        'lineup_hb_women_count': 'mujeres',
        'lineup_hb_gender_balance_error': 'Heartbreaker necesita exactamente 2 hombres y 2 mujeres',
        'lineup_mixed2_different_warning': 'Mixto 2 no puede incluir ningun jugador de Mixto 1',
        'upcoming_matches_heading': 'Proximos partidos',
        'past_matches_heading': 'Partidos pasados',
        'admin_override_mode': 'Modo Admin Override',
        'admin_edit_lineup': 'Ver/Editar alineacion',
        'admin_submit_for_team': 'Enviar alineacion para',
        'score_entry_title': 'Introducir resultado',
        'score_match_in_progress': 'Partido en curso',
        'score_match_completed': 'Partido finalizado',
        'score_heartbreaker_required': 'Heartbreaker requerido',
        'score_heartbreaker_section': 'Heartbreaker - 1 set a 21',
        'score_wins': 'gana',
        'score_home_team_score': 'Puntos local',
        'score_away_team_score': 'Puntos visitante',
        'score_save': 'Guardar',
        'score_edit': 'Editar',
        'score_refresh_standing': 'Actualizar marcador',
        'score_winner': 'Ganador',
        'standings_heading': 'Clasificacion',
        'standings_position': 'Pos',
        'standings_team': 'Equipo',
        'standings_played': 'J',
        'standings_wins': 'G',
        'standings_losses': 'P',
        'standings_points': 'Pts',
        'standings_diff': 'Diff',
        'standings_no_matches_yet': 'Sin partidos completados',
        'your_team_indicator': 'Tu equipo',
        'captain_pool_section': 'Jugadores del pool en',
        'profile': 'Perfil',
        'photo': 'Foto de perfil',
        'photo_help': 'Obligatorio. JPG, PNG, mÃƒÂ¡x 5MB. Formato cuadrado recomendado.',
        'upload_photo': 'Subir foto',
        'bio': 'Biografia breve',
        'bio_placeholder': 'Cuentanos sobre ti y tu viaje en pickleball...',
        'bio_help': 'Cuentanos sobre ti (50-500 caracteres)',
        'social_media': 'Redes Sociales',
        'optional_info': 'InformaciÃƒÂ³n Opcional',
        'video_url': 'URL del Video (Highlights)',
        'dupr_rating': 'Rating DUPR',
        'language': 'Idioma preferido',
        'privacy_accept': 'Acepto el procesamiento de datos para el registro PCL',
        'submit': 'Registrarse',
        'update': 'Actualizar registro',
        'required': 'Obligatorio',
        'optional': 'Opcional',
        'select': 'Seleccionar',
        'success_title': 'Registro completado!',
        'success_message': 'Gracias por registrarte en PCL.',
        'registration_whatsapp_sent': 'Registro exitoso! Revisa WhatsApp para tu enlace de perfil personal.',
        'captain_is_playing_label': 'El capitan tambien juega',
        'missing_fields': 'Por favor completa todos los campos obligatorios',
        'captain_dashboard': 'Panel del CapitÃƒÂ¡n',
        'team_status': 'Estado del Equipo',
        'registration_link': 'Enlace de registro para jugadores',
        'copy_link': 'Copiar enlace',
        'link_copied': 'Enlace copiado!',
        'players_registered': 'Jugadores',
        'men': 'Hombres',
        'women': 'Mujeres',
        'complete': 'Completo',
        'incomplete': 'Incompleto',
        'photo_missing': 'Falta foto',
        'deadline': 'Fecha limite',
        'days_left': 'dias restantes',
        'send_reminder': 'Enviar recordatorio',
        'export_data': 'Exportar datos del equipo',
        'missing': 'Falta',
        'no_players': 'Sin jugadores todavia',
        'requirements': 'Requisitos',
        'team_size': 'TamaÃƒÂ±o del equipo',
        'required_per_player': 'Requerido por jugador',
        'captain_link_warning': 'Este enlace es solo para ti como capitÃƒÂ¡n. No compartir!',
        # Quick Add specific
        'quick_add_title': 'AÃƒÂ±adir jugador rÃƒÂ¡pido',
        'quick_add_player': 'AÃƒÂ±adir jugador',
        'quick_add_info': 'El jugador completarÃƒÂ¡ su perfil a traves del enlace de WhatsApp',
        'add_player': 'AÃƒÂ±adir jugador',
        'back_to_dashboard': 'Volver al panel',
        'send_link_now': 'Enviar enlace WhatsApp ahora',
        'send_link_help': 'El jugador recibirÃƒÂ¡ un enlace para completar su perfil',
        'phone_invalid': 'Por favor ingresa un nÃƒÂºmero vÃƒÂ¡lido con cÃƒÂ³digo de pais',
        'quick_actions': 'Acciones rÃƒÂ¡pidas',
        'send_all_links': 'Enviar todos los enlaces',
        'send_all_confirm': 'Ã‚Â¿Enviar enlaces a todos los jugadores incompletos?',
        'player_added': 'Jugador aÃƒÂ±adido exitosamente!',
        'player_added_whatsapp': 'Jugador aÃƒÂ±adido y WhatsApp enviado!',
        'whatsapp_sent': 'WhatsApp enviado!',
        'whatsapp_failed': 'No se pudo enviar WhatsApp',
        # Complete Profile specific
        'complete_profile_title': 'Completa tu perfil',
        'hello': 'Hola',
        'profile_progress': 'Progreso del perfil',
        'save_profile': 'Guardar perfil',
        'profile_saved': 'Perfil guardado exitosamente!',
        'profile_updated': 'Perfil actualizado!',
        'edit_profile_title': 'Editar perfil',
        'current_photo': 'Foto actual',
        'leave_empty_to_keep': 'Dejar vacio para mantener',
        'data_protection': 'Tus datos estÃƒÂ¡n almacenados de forma segura',
        'contact': 'Contacto',
        'lock_closed_banner': 'Registro cerrado para este equipo. Contacta a Sergio para solicitar cambios.',
        'lock_open_until_banner': 'Registro abierto hasta {date}',
        'lock_open_manual_banner': 'Registro abierto (desbloqueado manualmente)',
        'lock_disabled_button': 'Bloqueado',
        'lock_admin_section': 'Bloqueo de registro',
        'lock_admin_toggle': 'Registro abierto (manual)',
        'lock_admin_deadline': 'Registro abierto hasta',
        'lock_admin_helper': 'Cuando esta marcado O la fecha es futura, el capitan puede anadir/editar/eliminar jugadores.',
    },
    'FR': {
        'page_title': 'Inscription Joueur PCL',
        'team': 'Ãƒâ€°quipe',
        'personal_info': 'Informations Personnelles',
        'first_name': 'Prenom',
        'last_name': 'Nom',
        'email': 'E-mail',
        'phone': 'Telephone',
        'phone_help': 'Avec indicatif pays (+33, +49, etc.)',
        'gender': 'Genre',
        'male': 'Homme',
        'female': 'Femme',
        'birth_year': 'Annee de naissance',
        'role': 'RÃƒÂ´le',
        'player': 'Joueur',
        'captain': 'Capitaine',
        'shirt_info': 'Informations Maillot',
        'shirt_size': 'Taille du maillot',
        'shirt_second': 'Second maillot (optionnel)',
        'shirt_third': 'TroisiÃ¨me maillot (optionnel)',
        'shirt_total_price': 'Total',
        'shirt_payment_info': 'Chaque maillot coÃ»te 15 EUR. Les maillots sont commandÃ©s et imprimÃ©s par nous selon les spÃ©cifications officielles de la PCL. Le paiement est gÃ©rÃ© par le capitaine de l\'Ã©quipe.',
        'shirt_product_preview_helper': "Voici le maillot que vous recevrez. Choisissez votre taille ci-dessous.",
        'match_status_pending': 'En attente',
        'match_status_lineups_locked': 'Compositions verrouillees',
        'match_status_in_progress': 'En cours',
        'match_status_completed': 'Termine',
        'match_create_new': 'Nouveau match',
        'match_team_home': 'Domicile',
        'match_team_away': 'Exterieur',
        'match_court': 'Court',
        'match_lineup_deadline': 'Date limite composition',
        'match_lineups_submitted': 'Compositions soumises',
        'match_delete_confirm': 'Supprimer le match?',
        'lineup_submit_title': 'Envoyer la composition',
        'lineup_edit_title': 'Modifier la composition',
        'lineup_section_wd': 'Double dames',
        'lineup_section_md': 'Double messieurs',
        'lineup_section_mx1': 'Mixte 1',
        'lineup_section_mx2': 'Mixte 2',
        'lineup_section_hb': 'Ordre Heartbreaker',
        'lineup_select_male': 'Choisir homme',
        'lineup_select_female': 'Choisir femme',
        'lineup_position': 'Position',
        'lineup_status_selected': 'selectionne',
        'lineup_submit_button': 'Envoyer la composition',
        'lineup_update_button': 'Mettre a jour',
        'lineup_already_submitted': 'Composition envoyee',
        'lineup_opponent_submitted': 'Adversaire a envoye',
        'lineup_opponent_pending': 'Adversaire pas encore envoye',
        'lineup_deadline_in': 'Date limite dans',
        'lineup_deadline_passed': 'Date limite passee',
        'lineup_reveal_title': 'Compositions revelees',
        'lineup_heartbreaker_help': '2 hommes + 2 femmes dans l ordre. Utilise si le match atteint 2:2.',
        'lineup_hb_men_count': 'hommes',
        'lineup_hb_women_count': 'femmes',
        'lineup_hb_gender_balance_error': 'Heartbreaker requiert exactement 2 hommes et 2 femmes',
        'lineup_mixed2_different_warning': 'Mixte 2 ne peut pas inclure de joueur de Mixte 1',
        'upcoming_matches_heading': 'Matchs a venir',
        'past_matches_heading': 'Matchs passes',
        'admin_override_mode': 'Mode Admin Override',
        'admin_edit_lineup': 'Voir/Modifier la composition',
        'admin_submit_for_team': 'Envoyer la composition pour',
        'score_entry_title': 'Saisir le resultat',
        'score_match_in_progress': 'Match en cours',
        'score_match_completed': 'Match termine',
        'score_heartbreaker_required': 'Heartbreaker requis',
        'score_heartbreaker_section': 'Heartbreaker - 1 set a 21',
        'score_wins': 'gagne',
        'score_home_team_score': 'Points domicile',
        'score_away_team_score': 'Points exterieur',
        'score_save': 'Enregistrer',
        'score_edit': 'Modifier',
        'score_refresh_standing': 'Actualiser le score',
        'score_winner': 'Vainqueur',
        'standings_heading': 'Classement',
        'standings_position': 'Pos',
        'standings_team': 'Equipe',
        'standings_played': 'J',
        'standings_wins': 'V',
        'standings_losses': 'D',
        'standings_points': 'Pts',
        'standings_diff': 'Diff',
        'standings_no_matches_yet': 'Aucun match termine',
        'your_team_indicator': 'Votre equipe',
        'captain_pool_section': 'Joueurs du pool en',
        'profile': 'Profil',
        'photo': 'Photo de profil',
        'photo_help': 'Obligatoire. JPG, PNG, max 5Mo. Format carre recommande.',
        'upload_photo': 'Telecharger photo',
        'bio': 'Courte bio',
        'bio_placeholder': 'Parlez-nous de vous et de votre parcours pickleball...',
        'bio_help': 'Parlez-nous de vous (50-500 caractÃƒÂ¨res)',
        'social_media': 'Reseaux Sociaux',
        'optional_info': 'Informations Optionnelles',
        'video_url': 'URL Video (Highlights)',
        'dupr_rating': 'Rating DUPR',
        'language': 'Langue preferee',
        'privacy_accept': "J'accepte le traitement des donnees pour l'inscription PCL",
        'submit': "S'inscrire",
        'update': "Mettre a jour l'inscription",
        'required': 'Obligatoire',
        'optional': 'Optionnel',
        'select': 'Selectionner',
        'success_title': 'Inscription reussie!',
        'success_message': 'Merci pour votre inscription a PCL.',
        'registration_whatsapp_sent': 'Inscription reussie! Consultez WhatsApp pour votre lien de profil personnel.',
        'captain_is_playing_label': 'Le capitaine joue aussi',
        'missing_fields': 'Veuillez remplir tous les champs obligatoires',
        'captain_dashboard': 'Tableau de bord Capitaine',
        'team_status': "Statut de l'equipe",
        'registration_link': "Lien d'inscription pour les joueurs",
        'copy_link': 'Copier le lien',
        'link_copied': 'Lien copie!',
        'players_registered': 'Joueurs',
        'men': 'Hommes',
        'women': 'Femmes',
        'complete': 'Complet',
        'incomplete': 'Incomplet',
        'photo_missing': 'Photo manquante',
        'deadline': 'Date limite',
        'days_left': 'jours restants',
        'send_reminder': 'Envoyer un rappel',
        'export_data': "Exporter les donnees de l'equipe",
        'missing': 'Manquant',
        'no_players': 'Pas encore de joueurs',
        'requirements': 'Exigences',
        'team_size': "Taille de l'equipe",
        'required_per_player': 'Requis par joueur',
        'captain_link_warning': 'Ce lien est uniquement pour vous en tant que capitaine. Ne pas partager!',
        # Quick Add specific
        'quick_add_title': 'Ajouter un joueur rapidement',
        'quick_add_player': 'Ajouter un joueur',
        'quick_add_info': 'Le joueur completera son profil via le lien WhatsApp',
        'add_player': 'Ajouter joueur',
        'back_to_dashboard': 'Retour au tableau de bord',
        'send_link_now': 'Envoyer le lien WhatsApp maintenant',
        'send_link_help': 'Le joueur recevra un lien pour completer son profil',
        'phone_invalid': 'Veuillez entrer un numero valide avec indicatif pays',
        'quick_actions': 'Actions rapides',
        'send_all_links': 'Envoyer tous les liens',
        'send_all_confirm': 'Envoyer les liens a tous les joueurs incomplets?',
        'player_added': 'Joueur ajoute avec succÃƒÂ¨s!',
        'player_added_whatsapp': 'Joueur ajoute et WhatsApp envoye!',
        'whatsapp_sent': 'WhatsApp envoye!',
        'whatsapp_failed': "Impossible d'envoyer WhatsApp",
        # Complete Profile specific
        'complete_profile_title': 'Completez votre profil',
        'hello': 'Bonjour',
        'profile_progress': 'Progression du profil',
        'save_profile': 'Enregistrer le profil',
        'profile_saved': 'Profil enregistre avec succÃƒÂ¨s!',
        'profile_updated': 'Profil mis a jour!',
        'edit_profile_title': 'Modifier le profil',
        'current_photo': 'Photo actuelle',
        'leave_empty_to_keep': 'Laisser vide pour conserver',
        'data_protection': 'Vos donnees sont stockees en toute securite',
        'contact': 'Contact',
        'lock_closed_banner': 'Inscriptions fermees pour cette equipe. Contactez Sergio pour demander des changements.',
        'lock_open_until_banner': 'Inscriptions ouvertes jusqu au {date}',
        'lock_open_manual_banner': 'Inscriptions ouvertes (debloque manuellement)',
        'lock_disabled_button': 'Verrouille',
        'lock_admin_section': 'Verrouillage des inscriptions',
        'lock_admin_toggle': 'Inscriptions ouvertes (manuel)',
        'lock_admin_deadline': 'Inscriptions ouvertes jusqu au',
        'lock_admin_helper': 'Si coche OU si la date est dans le futur, le capitaine peut ajouter/modifier/supprimer des joueurs.',
    }
}

def get_translations(lang='EN'):
    """Get translations for a language, fallback to EN"""
    return TRANSLATIONS.get(lang, TRANSLATIONS['EN'])


def registration_locked_redirect(team, token, lang='EN'):
    """Server-side lock guard for captain-facing roster routes.

    Returns a redirect response (with a flash) when the team's registration is
    closed, otherwise None so the caller can proceed. Admin routes never call this.
    """
    if team.is_registration_open():
        return None
    t = get_translations(lang)
    flash(t['lock_closed_banner'], 'warning')
    return redirect(url_for('pcl.captain_dashboard', token=token, lang=lang))


# ============================================================================
# WHATSAPP MESSAGE TEMPLATES FOR PROFILE COMPLETION
# ============================================================================

def get_profile_completion_message(registration, profile_url, lang='EN'):
    """Get WhatsApp message for profile completion in the specified language"""
    
    team = registration.team
    tournament = team.tournament
    
    messages = {
        'EN': f"""Ã°Å¸Ââ€œ PCL {tournament.name}

Hi {registration.first_name}! Ã°Å¸â€˜â€¹

You've been added to Team {team.country_flag} {team.country_name} ({team.age_category}).

Please complete your profile:
Ã°Å¸â€˜â€° {profile_url}

Required:
Ã¢Å“â€œ Profile photo
Ã¢Å“â€œ Shirt size
Ã¢Å“â€œ Short bio

Ã¢ÂÂ° Deadline: {tournament.registration_deadline.strftime('%d.%m.%Y')}

Questions? Contact your team captain.

See you on the court! 
WPC Series Europe""",

        'DE': f"""Ã°Å¸Ââ€œ PCL {tournament.name}

Hallo {registration.first_name}! Ã°Å¸â€˜â€¹

Du wurdest zu Team {team.country_flag} {team.country_name} ({team.age_category}) hinzugefuegt.

Bitte vervollstaendige dein Profil:
Ã°Å¸â€˜â€° {profile_url}

Erforderlich:
Ã¢Å“â€œ Profilbild
Ã¢Å“â€œ Shirt-Name & Groesse
Ã¢Å“â€œ Kurze Bio

Ã¢ÂÂ° Deadline: {tournament.registration_deadline.strftime('%d.%m.%Y')}

Fragen? Kontaktiere deinen Team-Kapitaen.

Bis bald auf dem Court! 
WPC Series Europe""",

        'ES': f"""Ã°Å¸Ââ€œ PCL {tournament.name}

Hola {registration.first_name}! Ã°Å¸â€˜â€¹

Has sido aÃƒÂ±adido al equipo {team.country_flag} {team.country_name} ({team.age_category}).

Por favor completa tu perfil:
Ã°Å¸â€˜â€° {profile_url}

Requerido:
Ã¢Å“â€œ Foto de perfil
Ã¢Å“â€œ Talla de camiseta
Ã¢Å“â€œ Breve biografia

Ã¢ÂÂ° Fecha limite: {tournament.registration_deadline.strftime('%d.%m.%Y')}

Ã‚Â¿Preguntas? Contacta a tu capitÃƒÂ¡n.

Nos vemos en la cancha! 
WPC Series Europe""",

        'FR': f"""Ã°Å¸Ââ€œ PCL {tournament.name}

Bonjour {registration.first_name}! Ã°Å¸â€˜â€¹

Vous avez ete ajoute a l'equipe {team.country_flag} {team.country_name} ({team.age_category}).

Veuillez completer votre profil:
Ã°Å¸â€˜â€° {profile_url}

Requis:
Ã¢Å“â€œ Photo de profil
Ã¢Å“â€œ Taille du maillot
Ã¢Å“â€œ Courte bio

Ã¢ÂÂ° Date limite: {tournament.registration_deadline.strftime('%d.%m.%Y')}

Questions? Contactez votre capitaine.

Ãƒâ‚¬ bientÃƒÂ´t sur le court! 
WPC Series Europe"""
    }
    
    return messages.get(lang, messages['EN'])


# ============================================================================
# ADMIN ROUTES
# ============================================================================

@pcl.route('/admin')
def admin_dashboard():
    """PCL Admin Dashboard - Overview of all tournaments"""
    tournaments = PCLTournament.query.order_by(PCLTournament.start_date.desc()).all()
    return render_template('pcl/admin_dashboard.html', tournaments=tournaments)


@pcl.route('/admin/tournament/create', methods=['GET', 'POST'])
def create_tournament():
    """Create a new PCL tournament"""
    if request.method == 'POST':
        tournament = PCLTournament(
            name=request.form['name'],
            start_date=datetime.strptime(request.form['start_date'], '%Y-%m-%d').date(),
            end_date=datetime.strptime(request.form['end_date'], '%Y-%m-%d').date(),
            location=request.form['location'],
            description=request.form.get('description'),
            registration_deadline=datetime.strptime(request.form['registration_deadline'], '%Y-%m-%dT%H:%M')
        )
        
        try:
            db.session.add(tournament)
            db.session.commit()
            flash(f'Tournament "{tournament.name}" created!', 'success')
            return redirect(url_for('pcl.admin_tournament_detail', tournament_id=tournament.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    
    return render_template('pcl/admin_tournament_form.html', tournament=None)


@pcl.route('/admin/tournament/<int:tournament_id>')
def admin_tournament_detail(tournament_id):
    """Admin view of a tournament with all teams"""
    tournament = PCLTournament.query.get_or_404(tournament_id)
    
    teams_19 = tournament.teams.filter_by(age_category='+19').order_by(PCLTeam.country_name).all()
    teams_50 = tournament.teams.filter_by(age_category='+50').order_by(PCLTeam.country_name).all()
    
    return render_template('pcl/admin_tournament_detail.html', 
                         tournament=tournament,
                         teams_19=teams_19,
                         teams_50=teams_50,
                         now=datetime.utcnow())


@pcl.route('/media/<int:tournament_id>')
def media_page(tournament_id):
    """Public media page with all player photos and info for social media"""
    tournament = PCLTournament.query.get_or_404(tournament_id)
    
    teams_data = []
    total_players = 0
    players_with_photos = 0
    
    for team in tournament.teams.order_by(PCLTeam.country_name).all():
        players = team.registrations.all()
        total_players += len(players)
        players_with_photos += len([p for p in players if p.photo_filename])
        
        if players:
            teams_data.append({
                'team': team,
                'players': players
            })
    
    return render_template('pcl/media_page.html',
                         tournament=tournament,
                         teams_data=teams_data,
                         total_players=total_players,
                         players_with_photos=players_with_photos)


@pcl.route('/admin/tournament/<int:tournament_id>/add-team', methods=['GET', 'POST'])
def add_team(tournament_id):
    """Add a team to a tournament"""
    tournament = PCLTournament.query.get_or_404(tournament_id)
    
    if request.method == 'POST':
        country_code = request.form['country_code'].upper()
        
        team = PCLTeam(
            tournament_id=tournament.id,
            country_code=country_code,
            country_name=request.form['country_name'],
            country_flag=COUNTRY_FLAGS.get(country_code, 'ðŸ³ï¸'),
            age_category=request.form['age_category'],
            min_men=int(request.form.get('min_men', 2)),
            max_men=int(request.form.get('max_men', 4)),
            min_women=int(request.form.get('min_women', 2)),
            max_women=int(request.form.get('max_women', 4)),
            captain_token=PCLTeam.generate_token()
        )
        
        try:
            db.session.add(team)
            db.session.commit()
            flash(f'Team {team.country_flag} {team.country_name} {team.age_category} added!', 'success')
            return redirect(url_for('pcl.admin_tournament_detail', tournament_id=tournament.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    
    return render_template('pcl/admin_add_team.html', 
                         tournament=tournament,
                         country_flags=COUNTRY_FLAGS)


@pcl.route('/admin/team/<int:team_id>')
def admin_team_detail(team_id):
    """Admin view of a specific team with captain management"""
    team = PCLTeam.query.get_or_404(team_id)
    
    playing_filter = db.or_(PCLRegistration.is_playing == True, PCLRegistration.is_captain == False)
    men = team.registrations.filter_by(gender='male').filter(playing_filter).all()
    women = team.registrations.filter_by(gender='female').filter(playing_filter).all()
    captains = team.registrations.filter_by(is_captain=True).all()
    stats = team.get_stats()
    
    return render_template('pcl/admin_team_detail.html',
                         team=team,
                         men=men,
                         women=women,
                         captains=captains,
                         stats=stats)


# ============================================================================
# ADMIN: PER-TEAM REGISTRATION LOCK
# ============================================================================

@pcl.route('/admin/team/<int:team_id>/toggle-registration', methods=['POST'])
def admin_toggle_registration(team_id):
    """Admin flips the manual registration_open boolean for a team."""
    team = PCLTeam.query.get_or_404(team_id)
    team.registration_open = not team.registration_open
    try:
        db.session.commit()
        state = 'OPEN' if team.registration_open else 'CLOSED'
        flash(f'Registration manual override is now {state}.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    return redirect(url_for('pcl.admin_team_detail', team_id=team_id))


@pcl.route('/admin/team/<int:team_id>/set-deadline', methods=['POST'])
def admin_set_registration_deadline(team_id):
    """Admin sets registration_open_until from a datetime-local form input."""
    team = PCLTeam.query.get_or_404(team_id)
    raw = (request.form.get('registration_open_until') or '').strip()
    if not raw:
        flash('Please pick a date and time.', 'warning')
        return redirect(url_for('pcl.admin_team_detail', team_id=team_id))
    parsed = None
    # datetime-local sends "YYYY-MM-DDTHH:MM" (seconds optional)
    for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%dT%H:%M:%S'):
        try:
            parsed = datetime.strptime(raw, fmt)
            break
        except ValueError:
            continue
    if parsed is None:
        flash('Invalid date format.', 'danger')
        return redirect(url_for('pcl.admin_team_detail', team_id=team_id))
    team.registration_open_until = parsed
    try:
        db.session.commit()
        flash(f'Registration open until {parsed.strftime("%d.%m.%Y %H:%M")}.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    return redirect(url_for('pcl.admin_team_detail', team_id=team_id))


@pcl.route('/admin/team/<int:team_id>/clear-deadline', methods=['POST'])
def admin_clear_registration_deadline(team_id):
    """Admin clears the registration_open_until deadline (sets it to NULL)."""
    team = PCLTeam.query.get_or_404(team_id)
    team.registration_open_until = None
    try:
        db.session.commit()
        flash('Registration deadline cleared.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    return redirect(url_for('pcl.admin_team_detail', team_id=team_id))


@pcl.route('/admin/team/<int:team_id>/add-captain', methods=['POST'])
def add_captain(team_id):
    """Admin adds a captain to a team"""
    team = PCLTeam.query.get_or_404(team_id)
    
    first_name = request.form.get('first_name', '').strip()
    last_name = request.form.get('last_name', '').strip()
    phone = request.form.get('phone', '').strip()
    gender = request.form.get('gender', 'male')
    language = request.form.get('language', 'EN')
    is_playing = request.form.get('is_playing') == 'on'
    send_whatsapp = request.form.get('send_whatsapp') == 'on'
    
    if not first_name or not last_name or not phone:
        flash('First name, last name and phone are required!', 'danger')
        return redirect(url_for('pcl.admin_team_detail', team_id=team_id))
    
    # Normalize phone
    phone = phone.replace(' ', '').replace('-', '')
    if not phone.startswith('+'):
        phone = '+' + phone
    
    captain = PCLRegistration(
        team_id=team.id,
        first_name=first_name,
        last_name=last_name,
        phone=phone,
        gender=gender,
        is_captain=True,
        is_playing=is_playing,
        preferred_language=language,
        status='incomplete'
    )
    
    captain.generate_profile_token()
    
    try:
        db.session.add(captain)
        db.session.commit()
        
        if is_playing:
            flash(f'Captain {first_name} {last_name} added as player!', 'success')
        else:
            flash(f'Captain {first_name} {last_name} added (not playing)!', 'success')
        
        # Send WhatsApp: (1) captain invitation template, (2) player registration link to forward
        if send_whatsapp and phone:
            # MESSAGE 1: Captain invitation via approved Content Template
            result1 = send_captain_invitation_template(
                team=team,
                captain_name=first_name,
                captain_phone=phone,
                captain_token=team.captain_token,
                language=language
            )

            if result1.get('status') == 'sent':
                captain.whatsapp_sent_at = datetime.now()
                db.session.commit()
                flash(f'Captain invitation sent to {first_name}!', 'success')
            else:
                flash(f'Captain invitation failed: {result1.get("error", "Unknown")}', 'warning')

            # MESSAGE 2: Registration link for captain to forward to team group
            team_name = f"{team.country_flag} {team.country_name} {team.age_category}"
            registration_url = request.host_url.rstrip('/') + url_for('pcl.player_register', token=team.captain_token) + '?lang=' + language
            player_link_messages = {
                'EN': f"Hi {first_name}, here is the registration link for {team_name} players. Please forward this to your team WhatsApp group:\n\n{registration_url}\n\nPlayers will be asked to complete their profile (photo, bio, shirt size).",
                'DE': f"Hallo {first_name}, hier ist der Registrierungslink fuer die Spieler von {team_name}. Bitte leite ihn an die Team-WhatsApp-Gruppe weiter:\n\n{registration_url}\n\nDie Spieler werden gebeten, ihr Profil zu vervollstaendigen (Foto, Bio, Shirtgroesse).",
                'ES': f"Hola {first_name}, aqui esta el enlace de registro para los jugadores de {team_name}. Por favor reenvialo al grupo WhatsApp:\n\n{registration_url}\n\nLos jugadores deben completar su perfil (foto, bio, talla).",
                'FR': f"Bonjour {first_name}, voici le lien d inscription pour les joueurs de {team_name}. Transmettez-le au groupe WhatsApp:\n\n{registration_url}\n\nLes joueurs doivent completer leur profil (photo, bio, taille).",
            }
            player_link_message = player_link_messages.get(language.upper(), player_link_messages['EN'])
            result2 = send_whatsapp_message(phone, player_link_message)
            if result2.get('status') != 'sent':
                flash(f'Player-link message failed: {result2.get("error", "Unknown")}', 'warning')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('pcl.admin_team_detail', team_id=team_id))


@pcl.route('/admin/team/<int:team_id>/export')
def export_team_data(team_id):
    """Export team data as CSV"""
    team = PCLTeam.query.get_or_404(team_id)
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow([
        'First Name', 'Last Name', 'Email', 'Phone', 'Gender', 'Birth Year',
        'Captain', 'Shirt Size', 'Shirt Size 2', 'Shirt Size 3', 'Bio',
        'Instagram', 'TikTok', 'YouTube', 'Twitter', 'DUPR', 'Status', 'Photo URL'
    ])

    for reg in team.registrations.all():
        writer.writerow([
            reg.first_name, reg.last_name, reg.email or '', reg.phone or '', reg.gender, reg.birth_year or '',
            'Yes' if reg.is_captain else 'No', reg.shirt_size or '', reg.shirt_size_2 or '', reg.shirt_size_3 or '', reg.bio or '',
            reg.instagram or '', reg.tiktok or '', reg.youtube or '', reg.twitter or '',
            reg.dupr_rating or '', reg.status, reg.photo_filename or ''
        ])
    
    output.seek(0)
    
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'pcl_{team.country_code}_{team.age_category}_players.csv'
    )



SHIRT_EXPORT_TRANSLATIONS = {
    'DE': {
        'title': 'PCL SHIRT BESTELLUNG - {tournament_name}',
        'summary_header': 'ZUSAMMENFASSUNG NACH GRÖSSEN',
        'gender_total_header': 'GESAMTBESTELLUNG NACH GESCHLECHT',
        'product_info_header': 'PRODUKT INFO',
        'men_product': 'Herren-Shirt',
        'women_product': 'Damen-Shirt',
        'size': 'Größe',
        'count': 'Anzahl',
        'total': 'GESAMT',
        'men': 'Herren',
        'women': 'Damen',
        'gender': 'Geschlecht',
        'no_gender': 'Ohne Geschlecht',
        'shirt_count_combined': 'Gesamt',
        'missing_data_header': 'FEHLENDE DATEN ({count} Spieler)',
        'player': 'Spieler',
        'team': 'Team',
        'category': 'Kategorie',
        'category_short': 'Kat.',
        'missing': 'Fehlt',
        'shirt_num': 'Shirt #',
        'size_section_header': 'GRÖSSE {size} ({count} Stück)',
        'size_section_header_singular': 'GRÖSSE {size} (1 Stück)',
        'size_gender_section_header': 'GRÖSSE {size} - {gender} ({count} Stück)',
        'size_gender_section_header_singular': 'GRÖSSE {size} - {gender} (1 Stück)',
        'sheet_overview': 'Bestellübersicht',
        'sheet_by_size': 'Nach Größe',
        'sheet_by_team': 'Nach Team',
        'filename_prefix': 'pcl_shirt_bestellung',
    },
    'ES': {
        'title': 'PEDIDO DE CAMISETAS PCL - {tournament_name}',
        'summary_header': 'RESUMEN POR TALLAS',
        'gender_total_header': 'PEDIDO TOTAL POR GÉNERO',
        'product_info_header': 'INFORMACIÓN DEL PRODUCTO',
        'men_product': 'Camiseta Hombre',
        'women_product': 'Camiseta Mujer',
        'size': 'Talla',
        'count': 'Cantidad',
        'total': 'TOTAL',
        'men': 'Hombre',
        'women': 'Mujer',
        'gender': 'Género',
        'no_gender': 'Sin género',
        'shirt_count_combined': 'Total',
        'missing_data_header': 'DATOS FALTANTES ({count} jugadores)',
        'player': 'Jugador',
        'team': 'Equipo',
        'category': 'Categoría',
        'category_short': 'Cat.',
        'missing': 'Falta',
        'shirt_num': 'Camiseta #',
        'size_section_header': 'TALLA {size} ({count} unidades)',
        'size_section_header_singular': 'TALLA {size} (1 unidad)',
        'size_gender_section_header': 'TALLA {size} - {gender} ({count} unidades)',
        'size_gender_section_header_singular': 'TALLA {size} - {gender} (1 unidad)',
        'sheet_overview': 'Resumen Pedido',
        'sheet_by_size': 'Por Talla',
        'sheet_by_team': 'Por Equipo',
        'filename_prefix': 'pedido_camisetas_pcl',
    },
}


@pcl.route('/admin/tournament/<int:tournament_id>/shirt-products', methods=['POST'])
def update_shirt_products(tournament_id):
    """Admin: upload men's/women's shirt product image + name for a tournament."""
    tournament = PCLTournament.query.get_or_404(tournament_id)

    # Men's product image
    if 'men_shirt_file' in request.files:
        f = request.files['men_shirt_file']
        if f and f.filename:
            result = upload_photo_to_supabase(f, folder='tournaments')
            if result['success']:
                tournament.men_shirt_image = result['url']
            else:
                flash(f"Men's shirt image upload failed: {result['error']}", 'warning')

    # Women's product image
    if 'women_shirt_file' in request.files:
        f = request.files['women_shirt_file']
        if f and f.filename:
            result = upload_photo_to_supabase(f, folder='tournaments')
            if result['success']:
                tournament.women_shirt_image = result['url']
            else:
                flash(f"Women's shirt image upload failed: {result['error']}", 'warning')

    # Product names (empty input clears the field)
    if 'men_shirt_name' in request.form:
        tournament.men_shirt_name = request.form.get('men_shirt_name', '').strip() or None
    if 'women_shirt_name' in request.form:
        tournament.women_shirt_name = request.form.get('women_shirt_name', '').strip() or None

    try:
        db.session.commit()
        flash('Shirt products updated.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')

    return redirect(url_for('pcl.admin_tournament_detail', tournament_id=tournament.id))


@pcl.route('/admin/export-shirts/<int:tournament_id>')
def export_shirt_list(tournament_id):
    """Export shirt list as Excel, split by gender, with product info (DE/ES via ?lang=)"""
    import tempfile
    import requests as http_requests
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage

    tournament = PCLTournament.query.get_or_404(tournament_id)

    # Language selection (default DE for backwards compatibility)
    lang = request.args.get('lang', 'DE').upper()
    if lang not in ['DE', 'ES']:
        lang = 'DE'
    t = SHIRT_EXPORT_TRANSLATIONS[lang]

    # Temp image files to embed; cleaned up after the workbook is saved.
    temp_image_files = []

    def embed_product_image(ws, anchor, url):
        """Download a Supabase image and embed it at the given cell. Fail soft."""
        if not url:
            return
        try:
            resp = http_requests.get(url, timeout=5)
            if resp.status_code != 200:
                print(f"Shirt product image fetch failed ({resp.status_code}): {url}")
                return
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.img')
            tmp.write(resp.content)
            tmp.close()
            temp_image_files.append(tmp.name)
            img = XLImage(tmp.name)
            img.width = 120
            img.height = 120
            ws.add_image(img, anchor)
        except Exception as e:
            print(f"Shirt product image embed failed for {url}: {e}")

    def gender_key(value):
        """Normalize the raw gender field to one of: male, female, unknown."""
        v = (value or '').strip().lower()
        if v == 'male':
            return 'male'
        if v == 'female':
            return 'female'
        return 'unknown'

    def gender_label(value):
        """Human label for a gender in the chosen language."""
        k = gender_key(value)
        if k == 'male':
            return t['men']
        if k == 'female':
            return t['women']
        return t['no_gender']

    # Collect all registrations (one entry per ordered shirt)
    all_regs = []
    incomplete = []

    for team in tournament.teams.all():
        for reg in team.registrations.all():
            base = {
                'team': f"{team.country_flag} {team.country_name}",
                'category': team.age_category,
                'player': f"{reg.first_name} {reg.last_name}",
                'gender': reg.gender,
            }
            size1 = (reg.shirt_size or '').strip()
            size2 = (reg.shirt_size_2 or '').strip()
            size3 = (reg.shirt_size_3 or '').strip()

            if size1:
                all_regs.append({**base, 'size': size1, 'shirt_no': 1})
            else:
                incomplete.append({**base, 'size': '', 'shirt_no': 1})

            if size2:
                all_regs.append({**base, 'size': size2, 'shirt_no': 2})

            if size3:
                all_regs.append({**base, 'size': size3, 'shirt_no': 3})

    # Shirts whose owner has an unknown gender (valid size, but no SKU side)
    no_gender_regs = [r for r in all_regs if gender_key(r['gender']) == 'unknown']

    # Create workbook
    wb = Workbook()

    # Styles
    header_fill = PatternFill('solid', fgColor='1F4E79')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    warning_fill = PatternFill('solid', fgColor='FFC7CE')
    warning_font = Font(bold=True, color='9C0006')
    yellow_fill = PatternFill('solid', fgColor='FFFFCC')
    size_header_fill = PatternFill('solid', fgColor='D9E1F2')
    border = Side(style='thin', color='000000')
    thin_border = Border(left=border, right=border, top=border, bottom=border)

    size_order = ['XS', 'S', 'M', 'L', 'XL', 'XXL', 'XXXL']

    # === SHEET 1: Order overview (by size and gender) ===
    ws1 = wb.active
    ws1.title = t['sheet_overview']

    # --- Optional product info block at the very top (rows 1-4) ---
    has_products = bool(
        tournament.men_shirt_name or tournament.men_shirt_image
        or tournament.women_shirt_name or tournament.women_shirt_image
    )
    offset = 0
    if has_products:
        ws1.cell(row=1, column=1, value=t['product_info_header'])
        ws1['A1'].font = Font(bold=True, size=12)
        ws1['A1'].fill = size_header_fill
        ws1.merge_cells('A1:D1')

        # Men row 2
        men_text = f"{t['men_product']}: {tournament.men_shirt_name or '-'}"
        ws1.cell(row=2, column=2, value=men_text)
        ws1['B2'].font = Font(bold=True)
        ws1['B2'].alignment = Alignment(vertical='center')
        ws1.merge_cells('B2:D2')
        embed_product_image(ws1, 'A2', tournament.men_shirt_image)

        # Women row 3
        women_text = f"{t['women_product']}: {tournament.women_shirt_name or '-'}"
        ws1.cell(row=3, column=2, value=women_text)
        ws1['B3'].font = Font(bold=True)
        ws1['B3'].alignment = Alignment(vertical='center')
        ws1.merge_cells('B3:D3')
        embed_product_image(ws1, 'A3', tournament.women_shirt_image)

        ws1.row_dimensions[2].height = 95
        ws1.row_dimensions[3].height = 95
        offset = 4  # row 4 acts as a spacer; main content starts at row 5

    # Title
    title_row = offset + 1
    ws1.cell(row=title_row, column=1, value=t['title'].format(tournament_name=tournament.name))
    ws1.cell(row=title_row, column=1).font = Font(bold=True, size=16)
    ws1.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=4)

    header_row3 = offset + 3
    ws1.cell(row=header_row3, column=1, value=t['gender_total_header'])
    ws1.cell(row=header_row3, column=1).font = Font(bold=True, size=14)

    # Per-size gender counts
    size_gender_counts = {}
    for reg in all_regs:
        size = reg['size'].upper()
        k = gender_key(reg['gender'])
        bucket = size_gender_counts.setdefault(size, {'male': 0, 'female': 0, 'unknown': 0})
        bucket[k] += 1

    # Header row: Size | Men | Women | Total
    table_header_row = offset + 4
    for col, label in enumerate([t['size'], t['men'], t['women'], t['shirt_count_combined']], 1):
        cell = ws1.cell(row=table_header_row, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row = offset + 5
    total_men = total_women = total_all = 0
    for size in size_order:
        bucket = size_gender_counts.get(size)
        if not bucket:
            continue
        men = bucket['male']
        women = bucket['female']
        line_total = men + women + bucket['unknown']
        if line_total == 0:
            continue
        total_men += men
        total_women += women
        total_all += line_total

        ws1.cell(row=row, column=1, value=size).border = thin_border
        for col, val in enumerate([men, women, line_total], 2):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        row += 1

    # Grand total row
    ws1.cell(row=row, column=1, value=t['total']).font = Font(bold=True)
    ws1.cell(row=row, column=1).border = thin_border
    for col, val in enumerate([total_men, total_women, total_all], 2):
        cell = ws1.cell(row=row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # Incomplete section (missing shirt size) - now with gender column
    if incomplete:
        row += 3
        ws1.cell(row=row, column=1, value=t['missing_data_header'].format(count=len(incomplete)))
        ws1.cell(row=row, column=1).font = warning_font
        ws1.cell(row=row, column=1).fill = warning_fill
        ws1.merge_cells(f'A{row}:E{row}')

        row += 1
        for col, header in enumerate([t['player'], t['team'], t['category'], t['gender'], t['missing']], 1):
            cell = ws1.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = thin_border

        row += 1
        for inc in incomplete:
            missing = []
            if not inc['size']:
                missing.append(t['size'])

            ws1.cell(row=row, column=1, value=inc['player']).border = thin_border
            ws1.cell(row=row, column=2, value=inc['team']).border = thin_border
            ws1.cell(row=row, column=3, value=inc['category']).border = thin_border
            ws1.cell(row=row, column=4, value=gender_label(inc['gender'])).border = thin_border
            ws1.cell(row=row, column=5, value=', '.join(missing)).border = thin_border
            for col in range(1, 6):
                ws1.cell(row=row, column=col).fill = yellow_fill
            row += 1

    # Players with a valid shirt size but unknown gender (needs SKU assignment)
    if no_gender_regs:
        row += 2
        ws1.cell(row=row, column=1, value=t['no_gender'])
        ws1.cell(row=row, column=1).font = warning_font
        ws1.cell(row=row, column=1).fill = warning_fill
        ws1.merge_cells(f'A{row}:E{row}')

        row += 1
        for col, header in enumerate([t['player'], t['team'], t['category'], t['size'], t['shirt_num']], 1):
            cell = ws1.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = thin_border

        row += 1
        for ng in sorted(no_gender_regs, key=lambda x: (x['size'], x['team'], x['player'], x['shirt_no'])):
            ws1.cell(row=row, column=1, value=ng['player']).border = thin_border
            ws1.cell(row=row, column=2, value=ng['team']).border = thin_border
            ws1.cell(row=row, column=3, value=ng['category']).border = thin_border
            ws1.cell(row=row, column=4, value=ng['size']).border = thin_border
            ws1.cell(row=row, column=5, value=ng['shirt_no']).border = thin_border
            for col in range(1, 6):
                ws1.cell(row=row, column=col).fill = yellow_fill
            row += 1

    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 20

    # === SHEET 2: Sorted by size, split by gender ===
    ws2 = wb.create_sheet(t['sheet_by_size'])
    row = 1

    gender_sections = [('male', t['men']), ('female', t['women']), ('unknown', t['no_gender'])]

    for size in size_order:
        size_regs = [r for r in all_regs if r['size'].upper() == size]
        if not size_regs:
            continue

        for gkey, glabel in gender_sections:
            sub = [r for r in size_regs if gender_key(r['gender']) == gkey]
            if not sub:
                continue

            if len(sub) == 1:
                header_text = t['size_gender_section_header_singular'].format(
                    size=size, gender=glabel.upper())
            else:
                header_text = t['size_gender_section_header'].format(
                    size=size, gender=glabel.upper(), count=len(sub))
            ws2[f'A{row}'] = header_text
            ws2[f'A{row}'].font = Font(bold=True, size=12)
            ws2[f'A{row}'].fill = size_header_fill
            ws2.merge_cells(f'A{row}:D{row}')
            row += 1

            for col, header in enumerate([t['player'], t['shirt_num'], t['team'], t['category_short']], 1):
                cell = ws2.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.border = thin_border
            row += 1

            for reg in sorted(sub, key=lambda x: (x['team'], x['player'], x['shirt_no'])):
                ws2.cell(row=row, column=1, value=reg['player']).border = thin_border
                ws2.cell(row=row, column=2, value=reg['shirt_no']).border = thin_border
                ws2.cell(row=row, column=3, value=reg['team']).border = thin_border
                ws2.cell(row=row, column=4, value=reg['category']).border = thin_border
                row += 1
            row += 1

    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 8

    # === SHEET 3: Sorted by team (with gender column) ===
    ws3 = wb.create_sheet(t['sheet_by_team'])

    headers = [t['team'], t['category'], t['player'], t['gender'], t['size'], t['shirt_num']]
    for col, h in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    sorted_regs = sorted(all_regs, key=lambda x: (x['team'], x['category'], x['player'], x['shirt_no']))
    for r, reg in enumerate(sorted_regs, 2):
        ws3.cell(row=r, column=1, value=reg['team']).border = thin_border
        ws3.cell(row=r, column=2, value=reg['category']).border = thin_border
        ws3.cell(row=r, column=3, value=reg['player']).border = thin_border
        ws3.cell(row=r, column=4, value=gender_label(reg['gender'])).border = thin_border
        ws3.cell(row=r, column=5, value=reg['size']).border = thin_border
        ws3.cell(row=r, column=6, value=reg['shirt_no']).border = thin_border

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 10
    ws3.column_dimensions['C'].width = 25
    ws3.column_dimensions['D'].width = 12
    ws3.column_dimensions['E'].width = 10
    ws3.column_dimensions['F'].width = 30

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Clean up temp image files now that the workbook is written
    for path in temp_image_files:
        try:
            os.remove(path)
        except Exception:
            pass

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"{t['filename_prefix']}_{tournament.id}.xlsx"
    )



# ============================================================================
# CAPTAIN ROUTES (Secret Link)
# ============================================================================
@pcl.route('/team/<token>')
def captain_dashboard(token):
    """Captain dashboard - accessed via secret link, shows all teams for this captain"""
    team = PCLTeam.query.filter_by(captain_token=token).first_or_404()
    
    lang = request.args.get('lang', 'EN').upper()
    if lang not in TRANSLATIONS:
        lang = 'EN'
    
    t = get_translations(lang)
    
    # Find captain in this team (to get phone number)
    current_captain = team.registrations.filter_by(is_captain=True).first()
    
    # Find ALL teams where this person is captain (by phone number)
    other_teams = []
    if current_captain and current_captain.phone:
        # Find all captain registrations with same phone
        all_captain_regs = PCLRegistration.query.filter_by(
            phone=current_captain.phone,
            is_captain=True
        ).all()
        
        # Get unique teams (excluding current)
        for reg in all_captain_regs:
            if reg.team_id != team.id:
                other_team = PCLTeam.query.get(reg.team_id)
                if other_team and other_team not in other_teams:
                    other_teams.append(other_team)
    
    stats = team.get_stats()
    men = team.registrations.filter_by(gender='male').all()
    women = team.registrations.filter_by(gender='female').all()
    
    days_left = (team.tournament.registration_deadline - datetime.now()).days
    
    registration_url = request.host_url.rstrip('/') + url_for('pcl.player_register', token=token)
    quick_add_url = request.host_url.rstrip('/') + url_for('pcl.quick_add_player', token=token)
    
    # Share messages for WhatsApp
    share_messages = {
        'EN': f"""Hello Team!

Please complete your PCL profile for {team.country_name} {team.age_category} at {team.tournament.name}:

{registration_url}?lang=EN

Required:
- Profile photo
- Short bio
- Shirt size

Thank you!""",
        'DE': f"""Hallo Team!

Bitte vervollstaendigt euer PCL Profil fuer {team.country_name} {team.age_category} bei {team.tournament.name}:

{registration_url}?lang=DE

Benoetigt werden:
- Profilbild
- Kurze Bio
- Shirtgroesse

Danke!""",
        'ES': f"""Hola Equipo!

Por favor completa tu perfil PCL para {team.country_name} {team.age_category} en {team.tournament.name}:

{registration_url}?lang=ES

Requerido:
- Foto de perfil
- Biografia breve
- Talla de camiseta

Gracias!""",
        'FR': f"""Bonjour l'equipe!

Veuillez completer votre profil PCL pour {team.country_name} {team.age_category} a {team.tournament.name}:

{registration_url}?lang=FR

Requis:
- Photo de profil
- Courte bio
- Taille du maillot

Merci!"""
    }
    
    share_message = share_messages.get(lang, share_messages['EN'])
    share_message_encoded = quote(share_message)
    
    # Individual player message
    player_messages = {
        'EN': f"Hi! Please complete your PCL profile for {team.country_name} {team.age_category}: {registration_url}?lang=EN",
        'DE': f"Hallo! Bitte vervollstaendige dein PCL Profil fuer {team.country_name} {team.age_category}: {registration_url}?lang=DE",
        'ES': f"Hola! Por favor completa tu perfil PCL para {team.country_name} {team.age_category}: {registration_url}?lang=ES",
        'FR': f"Salut! Veuillez completer votre profil PCL pour {team.country_name} {team.age_category}: {registration_url}?lang=FR"
    }
    player_message_encoded = quote(player_messages.get(lang, player_messages['EN']))

    # --- Phase 2: matches for this team (upcoming + past) ---
    team_matches = PCLMatch.query.filter(
        db.or_(PCLMatch.team_home_id == team.id, PCLMatch.team_away_id == team.id)
    ).all()

    def _match_view(m):
        is_home = m.team_home_id == team.id
        opponent = m.team_away if is_home else m.team_home
        own = m.lineups.filter_by(team_id=team.id).first()
        opp_id = m.team_away_id if is_home else m.team_home_id
        opp = m.lineups.filter_by(team_id=opp_id).first()
        return {
            'match': m,
            'opponent': opponent,
            'is_home': is_home,
            'own_submitted': bool(own and own.is_submitted()),
            'opp_submitted': bool(opp and opp.is_submitted()),
            'deadline_passed': m.is_lineup_deadline_passed(),
        }

    upcoming_matches = [_match_view(m) for m in team_matches if m.status != 'completed']
    upcoming_matches.sort(key=lambda v: (v['match'].match_date is None,
                                         v['match'].match_date or datetime.max))
    past_matches = [_match_view(m) for m in team_matches if m.status == 'completed']
    past_matches.sort(key=lambda v: (v['match'].match_date or datetime.min), reverse=True)

    deadline_soon = any(
        v['match'].lineup_deadline and not v['deadline_passed'] and not v['own_submitted']
        and (v['match'].lineup_deadline - datetime.now()).total_seconds() < 24 * 3600
        for v in upcoming_matches
    )

    # Live tournament standings (reuses the same helper as the admin match list)
    standings = get_tournament_standings(team.tournament_id)

    # Pool players matching this team's country + age category (collapsible section)
    from routes.pool import status_label as pool_status_label, STATUS_COLORS as POOL_STATUS_COLORS
    pool_matches = PoolPlayer.query.filter_by(
        country_name=team.country_name, age_category=team.age_category
    ).order_by(PoolPlayer.created_at.desc()).all()
    pool_players = [{
        'p': pp,
        'status_label': pool_status_label(pp.status, lang),
        'status_color': POOL_STATUS_COLORS.get(pp.status, 'secondary'),
    } for pp in pool_matches]

    return render_template('pcl/captain_dashboard.html',
                         team=team,
                         other_teams=other_teams,
                         upcoming_matches=upcoming_matches,
                         past_matches=past_matches,
                         deadline_soon=deadline_soon,
                         standings=standings,
                         pool_players=pool_players,
                         stats=stats,
                         men=men,
                         women=women,
                         days_left=days_left,
                         registration_url=registration_url,
                         quick_add_url=quick_add_url,
                         share_message=share_message,
                         share_message_encoded=share_message_encoded,
                         player_message_encoded=player_message_encoded,
                         t=t,
                         current_lang=lang)


# ============================================================================
# QUICK ADD PLAYER (NEW!)
# ============================================================================

@pcl.route('/team/<token>/quick-add', methods=['GET', 'POST'])
def quick_add_player(token):
    """Quick add player - minimal form for captains"""
    team = PCLTeam.query.filter_by(captain_token=token).first_or_404()
    
    lang = request.args.get('lang', request.form.get('preferred_language', 'EN')).upper()
    if lang not in TRANSLATIONS:
        lang = 'EN'
    
    t = get_translations(lang)
    stats = team.get_stats()

    # Per-team registration lock (admin-controlled)
    locked = registration_locked_redirect(team, token, lang)
    if locked:
        return locked

    # Check deadline
    if datetime.now() > team.tournament.registration_deadline:
        flash('Registration is closed.', 'danger')
        return redirect(url_for('pcl.captain_dashboard', token=token, lang=lang))
    
    if request.method == 'POST':
        first_name = request.form.get('first_name', '').strip()
        last_name = request.form.get('last_name', '').strip()
        phone = request.form.get('phone', '').strip()
        gender = request.form.get('gender')
        preferred_language = request.form.get('preferred_language', lang)
        send_whatsapp = request.form.get('send_whatsapp') == 'on'
        
        # Validate
        if not first_name or not last_name or not phone or not gender:
            flash(t['missing_fields'], 'danger')
            return redirect(url_for('pcl.quick_add_player', token=token, lang=lang))
        
        # Normalize phone number
        phone = phone.replace(' ', '').replace('-', '')
        if not phone.startswith('+'):
            phone = '+' + phone
        
        # Create registration with minimal data
        registration = PCLRegistration(
            team_id=team.id,
            first_name=first_name,
            last_name=last_name,
            phone=phone,
            gender=gender,
            preferred_language=preferred_language,
            status='incomplete'
        )
        
        # Generate profile token
        registration.generate_profile_token()
        
        try:
            db.session.add(registration)
            db.session.commit()
            
            # Send WhatsApp if requested
            if send_whatsapp and phone:
                profile_url = request.host_url.rstrip('/') + url_for('pcl.complete_profile', profile_token=registration.profile_token)
                message = get_profile_completion_message(registration, profile_url, preferred_language)
                sponsor_block = get_whatsapp_sponsor_block(pcl_tournament_id=team.tournament_id, language=preferred_language)
                message += sponsor_block

                result = send_whatsapp_message(phone, message, test_mode=False)
                
                if result.get('status') in ['sent', 'queued']:
                    flash(t['player_added_whatsapp'], 'success')
                else:
                    flash(f"{t['player_added']} ({t['whatsapp_failed']})", 'warning')
            else:
                flash(t['player_added'], 'success')
            
            return redirect(url_for('pcl.captain_dashboard', token=token, lang=lang))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    
    return render_template('pcl/quick_add_player.html',
                         team=team,
                         stats=stats,
                         t=t,
                         current_lang=lang)


# ============================================================================
# COMPLETE PROFILE (PUBLIC - via profile_token)
# ============================================================================
@pcl.route('/complete/<profile_token>', methods=['GET', 'POST'])
def complete_profile(profile_token):
    """Public profile completion page - accessed via WhatsApp link"""
    registration = PCLRegistration.query.filter_by(profile_token=profile_token).first_or_404()
    team = registration.team
    
    lang = request.args.get('lang', registration.preferred_language or 'EN').upper()
    if lang not in TRANSLATIONS:
        lang = 'EN'
    
    t = get_translations(lang)
    
    # Calculate completion percentage
    missing_fields = registration.get_missing_fields()
    total_required = 3  # photo, shirt_size, bio
    completed = total_required - len(missing_fields)
    completion_percent = int((completed / total_required) * 100)
    
    # Parse existing additional photos
    additional_photos_list = []
    if registration.additional_photos:
        try:
            additional_photos_list = json.loads(registration.additional_photos)
        except:
            additional_photos_list = []

    # Was the profile already complete before this request? (drives edit vs. complete wording)
    was_complete = registration.status == 'complete'

    if request.method == 'POST':
        # Handle profile photo upload
        if 'photo' in request.files:
            file = request.files['photo']
            if file and file.filename and allowed_file(file.filename):
                result = upload_photo_to_supabase(file, folder='players')
                if result['success']:
                    registration.photo_filename = result['url']
                else:
                    flash(f'Photo upload failed: {result["error"]}', 'warning')
        
        # Handle photos to delete
        photos_to_delete = request.form.get('photos_to_delete', '')
        if photos_to_delete:
            urls_to_delete = photos_to_delete.split('|||')
            additional_photos_list = [url for url in additional_photos_list if url not in urls_to_delete]
        
        # Handle additional photos upload
        if 'additional_photos' in request.files:
            files = request.files.getlist('additional_photos')
            for file in files:
                if file and file.filename and allowed_file(file.filename):
                    if len(additional_photos_list) >= 5:
                        break  # Max 5 photos
                    result = upload_photo_to_supabase(file, folder='players/social')
                    if result['success']:
                        additional_photos_list.append(result['url'])
        
        # Save additional photos as JSON
        registration.additional_photos = json.dumps(additional_photos_list) if additional_photos_list else None
        
        # Update other fields
        registration.shirt_size = request.form.get('shirt_size') or registration.shirt_size
        registration.shirt_size_2 = request.form.get('shirt_size_2') or registration.shirt_size_2
        registration.shirt_size_3 = request.form.get('shirt_size_3') or registration.shirt_size_3
        registration.bio = request.form.get('bio', '').strip() or registration.bio
        registration.email = request.form.get('email', '').strip() or registration.email
        registration.birth_year = int(request.form['birth_year']) if request.form.get('birth_year') else registration.birth_year
        registration.instagram = request.form.get('instagram', '').strip().replace('@', '') or registration.instagram
        registration.tiktok = request.form.get('tiktok', '').strip().replace('@', '') or registration.tiktok
        registration.youtube = request.form.get('youtube', '').strip() or registration.youtube
        registration.twitter = request.form.get('twitter', '').strip().replace('@', '') or registration.twitter
        registration.dupr_rating = request.form.get('dupr_rating', '').strip() or registration.dupr_rating
        registration.video_url = request.form.get('video_url', '').strip() or registration.video_url
        
        # Check completeness
        registration.check_completeness()
        
        try:
            db.session.commit()
            # If the profile was already complete, this was an edit -> "updated"; otherwise "saved"
            flash(t['profile_updated'] if was_complete else t['profile_saved'], 'success')
            return redirect(url_for('pcl.complete_profile', profile_token=profile_token, lang=lang))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    
    return render_template('pcl/complete_profile.html',
                         registration=registration,
                         team=team,
                         shirt_sizes=SHIRT_SIZES,
                         missing_fields=missing_fields,
                         completion_percent=completion_percent,
                         additional_photos_list=additional_photos_list,
                         is_edit=was_complete,
                         t=t,
                         current_lang=lang)


# ============================================================================
# SEND PROFILE LINKS (WhatsApp)
# ============================================================================

@pcl.route('/team/<token>/send-link/<int:registration_id>', methods=['POST'])
def send_single_profile_link(token, registration_id):
    """Send profile completion link to a single player"""
    team = PCLTeam.query.filter_by(captain_token=token).first_or_404()
    registration = PCLRegistration.query.get_or_404(registration_id)
    
    # Verify registration belongs to this team
    if registration.team_id != team.id:
        flash('Invalid request!', 'danger')
        return redirect(url_for('pcl.captain_dashboard', token=token))
    
    lang = request.args.get('lang', registration.preferred_language or 'EN')
    t = get_translations(lang)
    
    if not registration.phone:
        flash(f'{registration.first_name}: No phone number!', 'warning')
        return redirect(url_for('pcl.captain_dashboard', token=token, lang=lang))
    
    # Generate token if not exists
    if not registration.profile_token:
        registration.generate_profile_token()
        db.session.commit()
    
    # Build URL and send
    profile_url = request.host_url.rstrip('/') + url_for('pcl.complete_profile', profile_token=registration.profile_token)
    message = get_profile_completion_message(registration, profile_url, registration.preferred_language or 'EN')
    sponsor_block = get_whatsapp_sponsor_block(pcl_tournament_id=team.tournament_id, language=registration.preferred_language or 'EN')
    message += sponsor_block

    result = send_whatsapp_message(registration.phone, message, test_mode=False)
    
    if result.get('status') in ['sent', 'queued']:
        registration.whatsapp_sent_at = datetime.now()
        db.session.commit()
        flash(f'{registration.first_name}: {t["whatsapp_sent"]}', 'success')
    else:
        flash(f'{registration.first_name}: {t["whatsapp_failed"]}', 'danger')
    
    return redirect(url_for('pcl.captain_dashboard', token=token, lang=lang))


@pcl.route('/team/<token>/send-all-links', methods=['POST'])
def send_all_profile_links(token):
    """Send profile completion links to all incomplete players"""
    team = PCLTeam.query.filter_by(captain_token=token).first_or_404()
    
    lang = request.args.get('lang', 'EN')
    t = get_translations(lang)
    
    sent_count = 0
    error_count = 0
    
    # Get all incomplete registrations with phone numbers
    incomplete = team.registrations.filter(
        PCLRegistration.status != 'complete',
        PCLRegistration.phone.isnot(None),
        PCLRegistration.phone != ''
    ).all()
    
    for registration in incomplete:
        # Generate token if not exists
        if not registration.profile_token:
            registration.generate_profile_token()
        
        profile_url = request.host_url.rstrip('/') + url_for('pcl.complete_profile', profile_token=registration.profile_token)
        message = get_profile_completion_message(registration, profile_url, registration.preferred_language or 'EN')
        sponsor_block = get_whatsapp_sponsor_block(pcl_tournament_id=team.tournament_id, language=registration.preferred_language or 'EN')
        message += sponsor_block

        result = send_whatsapp_message(registration.phone, message, test_mode=False)
        
        if result.get('status') in ['sent', 'queued']:
            sent_count += 1
        else:
            error_count += 1
    
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Error saving: {str(e)}', 'danger')
        return redirect(url_for('pcl.captain_dashboard', token=token, lang=lang))
    
    if sent_count > 0:
        flash(f'Ã¢Å“â€¦ {sent_count} {t["whatsapp_sent"]}', 'success')
    if error_count > 0:
        flash(f'Ã¢Å¡Â Ã¯Â¸Â {error_count} {t["whatsapp_failed"]}', 'warning')
    if sent_count == 0 and error_count == 0:
        flash('No incomplete players with phone numbers found.', 'info')
    
    return redirect(url_for('pcl.captain_dashboard', token=token, lang=lang))


# ============================================================================
# DELETE PLAYER
# ============================================================================

@pcl.route('/team/<token>/delete-player/<int:registration_id>', methods=['POST'])
def delete_player(token, registration_id):
    """Delete a player from the team (captain only)"""
    team = PCLTeam.query.filter_by(captain_token=token).first_or_404()
    registration = PCLRegistration.query.get_or_404(registration_id)

    lang = request.args.get('lang', 'EN')

    # Per-team registration lock (admin-controlled)
    locked = registration_locked_redirect(team, token, lang)
    if locked:
        return locked

    # Verify registration belongs to this team
    if registration.team_id != team.id:
        flash('Invalid request!', 'danger')
        return redirect(url_for('pcl.captain_dashboard', token=token))

    # Don't allow deleting captain
    if registration.is_captain:
        flash('Cannot delete the captain!', 'danger')
        return redirect(url_for('pcl.captain_dashboard', token=token))

    player_name = f"{registration.first_name} {registration.last_name}"
    
    try:
        db.session.delete(registration)
        db.session.commit()
        flash(f'{player_name} wurde geloescht.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('pcl.captain_dashboard', token=token, lang=lang))


# ============================================================================
# PLAYER REGISTRATION (Original Full Form)
# ============================================================================

@pcl.route('/register/<token>', methods=['GET', 'POST'])
def player_register(token):
    """Player registration form with Supabase photo upload"""
    team = PCLTeam.query.filter_by(captain_token=token).first_or_404()
    
    lang = request.args.get('lang', request.form.get('preferred_language', 'EN')).upper()
    if lang not in TRANSLATIONS:
        lang = 'EN'
    
    t = get_translations(lang)

    # Per-team registration lock (admin-controlled)
    locked = registration_locked_redirect(team, token, lang)
    if locked:
        return locked

    # Check if registration is still open
    if datetime.now() > team.tournament.registration_deadline:
        flash('Registration is closed.', 'danger')
        return redirect(url_for('pcl.captain_dashboard', token=token))
    
    if request.method == 'POST':
        # Handle profile photo upload to Supabase
        photo_url = None
        if 'photo' in request.files:
            file = request.files['photo']
            if file and file.filename and allowed_file(file.filename):
                result = upload_photo_to_supabase(file, folder='players')
                if result['success']:
                    photo_url = result['url']
                else:
                    flash(f'Photo upload failed: {result["error"]}', 'warning')
        
        # Handle additional photos upload
        additional_photos_list = []
        if 'additional_photos' in request.files:
            files = request.files.getlist('additional_photos')
            for file in files:
                if file and file.filename and allowed_file(file.filename):
                    if len(additional_photos_list) >= 5:
                        break  # Max 5 photos
                    result = upload_photo_to_supabase(file, folder='players/social')
                    if result['success']:
                        additional_photos_list.append(result['url'])
        
        # Create registration
        registration = PCLRegistration(
            team_id=team.id,
            first_name=request.form['first_name'],
            last_name=request.form['last_name'],
            email=request.form.get('email'),
            phone=request.form.get('phone'),
            gender=request.form['gender'],
            birth_year=int(request.form['birth_year']) if request.form.get('birth_year') else None,
            is_captain=request.form.get('is_captain') == 'on',
            shirt_size=request.form.get('shirt_size'),
            shirt_size_2=request.form.get('shirt_size_2') or None,
            shirt_size_3=request.form.get('shirt_size_3') or None,
            photo_filename=photo_url,
            bio=request.form.get('bio'),
            instagram=request.form.get('instagram', '').replace('@', ''),
            tiktok=request.form.get('tiktok', '').replace('@', ''),
            youtube=request.form.get('youtube'),
            twitter=request.form.get('twitter', '').replace('@', ''),
            video_url=request.form.get('video_url'),
            dupr_rating=request.form.get('dupr_rating'),
            preferred_language=lang,
            additional_photos=json.dumps(additional_photos_list) if additional_photos_list else None
        )
        
        registration.generate_profile_token()
        registration.check_completeness()
        
        try:
            db.session.add(registration)
            db.session.commit()

            # Send the player their personal profile link via WhatsApp so they can edit anytime.
            # A WhatsApp failure must NOT undo the already-committed registration.
            whatsapp_sent = False
            if registration.phone:
                try:
                    profile_url = request.host_url.rstrip('/') + url_for('pcl.complete_profile', profile_token=registration.profile_token)
                    message = get_profile_completion_message(registration, profile_url, registration.preferred_language or 'EN')
                    sponsor_block = get_whatsapp_sponsor_block(pcl_tournament_id=team.tournament_id, language=registration.preferred_language or 'EN')
                    message += sponsor_block

                    result = send_whatsapp_message(registration.phone, message, test_mode=False)

                    if result.get('status') in ['sent', 'queued']:
                        registration.whatsapp_sent_at = datetime.now()
                        db.session.commit()
                        whatsapp_sent = True
                except Exception as wa_error:
                    db.session.rollback()
                    print(f'WhatsApp send failed for registration {registration.id}: {wa_error}')

            if whatsapp_sent:
                flash(t['registration_whatsapp_sent'], 'success')
            else:
                flash(t['success_message'], 'success')

            return redirect(url_for('pcl.registration_success',
                                  registration_id=registration.id,
                                  lang=lang))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    
    return render_template('pcl/player_register.html',
                         team=team,
                         shirt_sizes=SHIRT_SIZES,
                         t=t,
                         current_lang=lang)



@pcl.route('/register/success/<int:registration_id>')
def registration_success(registration_id):
    """Registration success page"""
    registration = PCLRegistration.query.get_or_404(registration_id)
    
    lang = request.args.get('lang', registration.preferred_language).upper()
    t = get_translations(lang)
    
    return render_template('pcl/registration_success.html',
                         registration=registration,
                         t=t)


@pcl.route('/register/edit/<int:registration_id>', methods=['GET', 'POST'])
def edit_registration(registration_id):
    """Edit existing registration (captain) - pre-fills with the player's existing data"""
    registration = PCLRegistration.query.get_or_404(registration_id)
    team = registration.team

    lang = request.args.get('lang', registration.preferred_language or 'EN').upper()
    if lang not in TRANSLATIONS:
        lang = 'EN'
    t = get_translations(lang)

    # Per-team registration lock (admin-controlled)
    locked = registration_locked_redirect(team, team.captain_token, lang)
    if locked:
        return locked

    # Parse existing additional photos so the template can render them as thumbnails
    additional_photos_list = []
    if registration.additional_photos:
        try:
            additional_photos_list = json.loads(registration.additional_photos)
        except:
            additional_photos_list = []

    if request.method == 'POST':
        # Preserve-on-empty: only overwrite a field when a new value was submitted
        registration.first_name = request.form.get('first_name') or registration.first_name
        registration.last_name = request.form.get('last_name') or registration.last_name
        registration.email = request.form.get('email') or registration.email
        registration.phone = request.form.get('phone') or registration.phone
        registration.gender = request.form.get('gender') or registration.gender
        registration.birth_year = int(request.form['birth_year']) if request.form.get('birth_year') else registration.birth_year
        # is_captain has no control on this form -> preserve existing value
        if 'is_captain' in request.form:
            registration.is_captain = request.form.get('is_captain') == 'on'
        # is_playing checkbox is only rendered for captains in edit mode. The hidden
        # 'is_playing_present' marker tells us the checkbox was shown, so an unchecked
        # box (which browsers omit from the POST) correctly sets is_playing = False.
        if 'is_playing_present' in request.form:
            registration.is_playing = request.form.get('is_playing') == 'on'
        registration.shirt_size = request.form.get('shirt_size') or registration.shirt_size
        registration.shirt_size_2 = request.form.get('shirt_size_2') or registration.shirt_size_2
        registration.shirt_size_3 = request.form.get('shirt_size_3') or registration.shirt_size_3
        registration.bio = request.form.get('bio', '').strip() or registration.bio
        registration.instagram = request.form.get('instagram', '').strip().replace('@', '') or registration.instagram
        registration.tiktok = request.form.get('tiktok', '').strip().replace('@', '') or registration.tiktok
        registration.youtube = request.form.get('youtube', '').strip() or registration.youtube
        registration.twitter = request.form.get('twitter', '').strip().replace('@', '') or registration.twitter
        registration.video_url = request.form.get('video_url', '').strip() or registration.video_url
        registration.dupr_rating = request.form.get('dupr_rating', '').strip() or registration.dupr_rating
        registration.preferred_language = lang

        # Handle new profile photo upload (leave empty to keep current one)
        if 'photo' in request.files:
            file = request.files['photo']
            if file and file.filename and allowed_file(file.filename):
                result = upload_photo_to_supabase(file, folder='players')
                if result['success']:
                    registration.photo_filename = result['url']
                else:
                    flash(f'Photo upload failed: {result["error"]}', 'warning')

        # Handle additional photos to delete
        photos_to_delete = request.form.get('photos_to_delete', '')
        if photos_to_delete:
            urls_to_delete = photos_to_delete.split('|||')
            additional_photos_list = [url for url in additional_photos_list if url not in urls_to_delete]

        # Handle additional photos upload
        if 'additional_photos' in request.files:
            files = request.files.getlist('additional_photos')
            for file in files:
                if file and file.filename and allowed_file(file.filename):
                    if len(additional_photos_list) >= 5:
                        break  # Max 5 photos
                    result = upload_photo_to_supabase(file, folder='players/social')
                    if result['success']:
                        additional_photos_list.append(result['url'])

        registration.additional_photos = json.dumps(additional_photos_list) if additional_photos_list else None

        registration.check_completeness()

        try:
            db.session.commit()
            flash(t['profile_updated'], 'success')
            # Back to the captain dashboard so the captain sees the updated player in the team list
            return redirect(url_for('pcl.captain_dashboard', token=team.captain_token, lang=lang))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')

    return render_template('pcl/player_register.html',
                         team=team,
                         registration=registration,
                         shirt_sizes=SHIRT_SIZES,
                         additional_photos_list=additional_photos_list,
                         t=t,
                         current_lang=lang,
                         edit_mode=True)


# ============================================================================
# ADMIN REGISTRATION EDIT
# ============================================================================

@pcl.route('/admin/registration/<int:registration_id>/edit', methods=['GET', 'POST'])
def admin_edit_registration(registration_id):
    """Admin edit registration"""
    registration = PCLRegistration.query.get_or_404(registration_id)
    team = registration.team
    
    if request.method == 'POST':
        registration.first_name = request.form['first_name']
        registration.last_name = request.form['last_name']
        registration.email = request.form.get('email')
        registration.phone = request.form.get('phone')
        registration.gender = request.form['gender']
        registration.birth_year = int(request.form['birth_year']) if request.form.get('birth_year') else None
        registration.is_captain = request.form.get('is_captain') == 'on'
        registration.shirt_size = request.form.get('shirt_size')
        registration.shirt_size_2 = request.form.get('shirt_size_2') or None
        registration.shirt_size_3 = request.form.get('shirt_size_3') or None
        registration.bio = request.form.get('bio')
        registration.instagram = request.form.get('instagram', '').replace('@', '')
        registration.tiktok = request.form.get('tiktok', '').replace('@', '')
        registration.youtube = request.form.get('youtube')
        registration.twitter = request.form.get('twitter', '').replace('@', '')
        registration.video_url = request.form.get('video_url')
        registration.dupr_rating = request.form.get('dupr_rating')
        registration.preferred_language = request.form.get('preferred_language', 'EN')
        registration.status = request.form.get('status', registration.status)
        
        # Handle photo URL or file upload
        photo_url = request.form.get('photo_filename')
        if photo_url:
            registration.photo_filename = photo_url
        
        if 'photo' in request.files:
            file = request.files['photo']
            if file and file.filename and allowed_file(file.filename):
                result = upload_photo_to_supabase(file, folder='players')
                if result['success']:
                    registration.photo_filename = result['url']
        
        registration.check_completeness()
        
        try:
            db.session.commit()
            flash(f'Registration for {registration.first_name} {registration.last_name} updated!', 'success')
            return redirect(url_for('pcl.admin_team_detail', team_id=team.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    
    return render_template('pcl/pcl_registration_edit.html',
                         registration=registration,
                         team=team,
                         t=get_translations(registration.preferred_language or 'EN'))


# ============================================================================
# PCL MATCH / LINEUP MODULE - PHASE 1 (Admin CRUD)
# ============================================================================

_MATCH_WEEKDAYS = {
    'DE': ['Mo', 'Di', 'Mi', 'Do', 'Fr', 'Sa', 'So'],
    'EN': ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'],
    'ES': ['Lun', 'Mar', 'Mie', 'Jue', 'Vie', 'Sab', 'Dom'],
    'FR': ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim'],
}


def format_match_date(dt, lang='DE'):
    """Format a datetime as 'Mi 02.07. 10:00 Uhr' (DE) / 'Wed 02.07. 10:00' (other langs)."""
    if not dt:
        return ''
    lang = (lang or 'DE').upper()
    days = _MATCH_WEEKDAYS.get(lang, _MATCH_WEEKDAYS['DE'])
    weekday = days[dt.weekday()]
    base = dt.strftime(f'{weekday} %d.%m. %H:%M')
    return f'{base} Uhr' if lang == 'DE' else base


def _match_lang():
    """Resolve the request language for match admin pages (default EN)."""
    lang = request.args.get('lang', 'EN').upper()
    if lang not in TRANSLATIONS:
        lang = 'EN'
    return lang


def _parse_match_dt(value):
    """Parse a datetime-local form value ('YYYY-MM-DDTHH:MM') into a datetime or None."""
    if not value:
        return None
    for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%dT%H:%M:%S'):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def get_tournament_standings(tournament_id):
    """Live standings grouped by age category, computed from completed matches.

    1 point per win; sorted by points DESC, discipline differential DESC
    (sum of this team's score minus opponent's across completed matches),
    then country name ASC. Teams with no completed matches appear with 0 stats.
    """
    teams = PCLTeam.query.filter_by(tournament_id=tournament_id).all()
    completed = PCLMatch.query.filter_by(tournament_id=tournament_id, status='completed').all()

    stats = {team.id: {'team': team, 'played': 0, 'wins': 0, 'losses': 0, 'points': 0, 'diff': 0}
             for team in teams}

    for m in completed:
        hs = m.home_score or 0
        away = m.away_score or 0
        for tid, is_home in ((m.team_home_id, True), (m.team_away_id, False)):
            s = stats.get(tid)
            if s is None:
                continue
            s['played'] += 1
            s['diff'] += (hs - away) if is_home else (away - hs)
            if m.winner_id == tid:
                s['wins'] += 1
                s['points'] += 1
            else:
                s['losses'] += 1

    groups = {}
    for team in teams:
        groups.setdefault(team.age_category, []).append(stats[team.id])
    for cat in groups:
        groups[cat].sort(key=lambda s: (-s['points'], -s['diff'], (s['team'].country_name or '').lower()))

    # Stable category order (e.g. +19 before +50); skip nothing - empty cats never created.
    return {cat: groups[cat] for cat in sorted(groups)}


@pcl.route('/admin/tournament/<int:tournament_id>/matches')
def admin_match_list(tournament_id):
    """Admin: list all matches for a tournament, sorted by match date."""
    tournament = PCLTournament.query.get_or_404(tournament_id)
    lang = _match_lang()
    t = get_translations(lang)

    matches = PCLMatch.query.filter_by(tournament_id=tournament_id).all()
    matches.sort(key=lambda m: (m.match_date is None, m.match_date or datetime.max))

    standings = get_tournament_standings(tournament_id)

    return render_template('pcl/admin_match_list.html',
                         tournament=tournament,
                         matches=matches,
                         standings=standings,
                         t=t,
                         current_lang=lang)


@pcl.route('/admin/match/create/<int:tournament_id>', methods=['GET', 'POST'])
def admin_match_create(tournament_id):
    """Admin: create a new match for a tournament."""
    tournament = PCLTournament.query.get_or_404(tournament_id)
    lang = _match_lang()
    t = get_translations(lang)
    teams = tournament.teams.order_by(PCLTeam.country_name).all()

    if request.method == 'POST':
        team_home_id = request.form.get('team_home_id', type=int)
        team_away_id = request.form.get('team_away_id', type=int)
        match_date = _parse_match_dt(request.form.get('match_date'))
        lineup_deadline = _parse_match_dt(request.form.get('lineup_deadline'))
        court = (request.form.get('court') or '').strip() or None
        notes = (request.form.get('notes') or '').strip() or None

        errors = []
        team_home = PCLTeam.query.get(team_home_id) if team_home_id else None
        team_away = PCLTeam.query.get(team_away_id) if team_away_id else None

        if not team_home or not team_away:
            errors.append('Please choose both teams.')
        elif team_home_id == team_away_id:
            errors.append('Home and away team must be different.')
        if team_home and team_home.tournament_id != tournament.id:
            errors.append('Home team does not belong to this tournament.')
        if team_away and team_away.tournament_id != tournament.id:
            errors.append('Away team does not belong to this tournament.')
        if not match_date:
            errors.append('Please choose a valid match date and time.')
        elif match_date <= datetime.now():
            errors.append('Match date must be in the future.')
        if match_date and lineup_deadline and lineup_deadline >= match_date:
            errors.append('Lineup deadline must be before the match date.')

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('pcl/admin_match_form.html',
                                 tournament=tournament, teams=teams, match=None,
                                 form_data=request.form, t=t, current_lang=lang)

        match = PCLMatch(
            tournament_id=tournament.id,
            team_home_id=team_home_id,
            team_away_id=team_away_id,
            match_date=match_date,
            lineup_deadline=lineup_deadline,
            court=court,
            notes=notes,
            status='pending',
        )
        try:
            db.session.add(match)
            db.session.commit()
            flash('Match created.', 'success')
            return redirect(url_for('pcl.admin_match_detail', match_id=match.id, lang=lang))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')

    return render_template('pcl/admin_match_form.html',
                         tournament=tournament, teams=teams, match=None,
                         form_data={}, t=t, current_lang=lang)


@pcl.route('/admin/match/<int:match_id>')
def admin_match_detail(match_id):
    """Admin: detail view of a single match."""
    match = PCLMatch.query.get_or_404(match_id)
    lang = _match_lang()
    t = get_translations(lang)

    lineups = match.lineups.all()
    home_lineup = next((l for l in lineups if l.team_id == match.team_home_id), None)
    away_lineup = next((l for l in lineups if l.team_id == match.team_away_id), None)
    results = match.results.all()

    return render_template('pcl/admin_match_detail.html',
                         tournament=match.tournament,
                         match=match,
                         home_lineup=home_lineup,
                         away_lineup=away_lineup,
                         results=results,
                         t=t,
                         current_lang=lang)


@pcl.route('/admin/match/<int:match_id>/edit', methods=['GET', 'POST'])
def admin_match_edit(match_id):
    """Admin: edit an existing match."""
    match = PCLMatch.query.get_or_404(match_id)
    tournament = match.tournament
    lang = _match_lang()
    t = get_translations(lang)
    teams = tournament.teams.order_by(PCLTeam.country_name).all()

    if request.method == 'POST':
        team_home_id = request.form.get('team_home_id', type=int)
        team_away_id = request.form.get('team_away_id', type=int)
        match_date = _parse_match_dt(request.form.get('match_date'))
        lineup_deadline = _parse_match_dt(request.form.get('lineup_deadline'))
        court = (request.form.get('court') or '').strip() or None
        notes = (request.form.get('notes') or '').strip() or None
        status = request.form.get('status') or match.status

        errors = []
        team_home = PCLTeam.query.get(team_home_id) if team_home_id else None
        team_away = PCLTeam.query.get(team_away_id) if team_away_id else None

        if not team_home or not team_away:
            errors.append('Please choose both teams.')
        elif team_home_id == team_away_id:
            errors.append('Home and away team must be different.')
        if team_home and team_home.tournament_id != tournament.id:
            errors.append('Home team does not belong to this tournament.')
        if team_away and team_away.tournament_id != tournament.id:
            errors.append('Away team does not belong to this tournament.')
        if not match_date:
            errors.append('Please choose a valid match date and time.')
        if match_date and lineup_deadline and lineup_deadline >= match_date:
            errors.append('Lineup deadline must be before the match date.')

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('pcl/admin_match_form.html',
                                 tournament=tournament, teams=teams, match=match,
                                 form_data=request.form, t=t, current_lang=lang)

        match.team_home_id = team_home_id
        match.team_away_id = team_away_id
        match.match_date = match_date
        match.lineup_deadline = lineup_deadline
        match.court = court
        match.notes = notes
        if status in ('pending', 'lineups_locked', 'in_progress', 'completed'):
            match.status = status

        try:
            db.session.commit()
            flash('Match updated.', 'success')
            return redirect(url_for('pcl.admin_match_detail', match_id=match.id, lang=lang))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')

    return render_template('pcl/admin_match_form.html',
                         tournament=tournament, teams=teams, match=match,
                         form_data={}, t=t, current_lang=lang)


@pcl.route('/admin/match/<int:match_id>/delete', methods=['POST'])
def admin_match_delete(match_id):
    """Admin: delete a match (cascades to lineups + results)."""
    match = PCLMatch.query.get_or_404(match_id)
    tournament_id = match.tournament_id
    lang = _match_lang()
    try:
        db.session.delete(match)
        db.session.commit()
        flash('Match deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    return redirect(url_for('pcl.admin_match_list', tournament_id=tournament_id, lang=lang))


# ============================================================================
# DELETE ROUTES
# ============================================================================

@pcl.route('/admin/team/<int:team_id>/delete', methods=['POST'])
def delete_team(team_id):
    """Delete a team and all its registrations"""
    team = PCLTeam.query.get_or_404(team_id)
    tournament_id = team.tournament_id
    team_name = f"{team.country_name} {team.age_category}"
    
    try:
        PCLRegistration.query.filter_by(team_id=team_id).delete()
        db.session.delete(team)
        db.session.commit()
        flash(f'Team "{team_name}" deleted!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting team: {str(e)}', 'danger')
    
    return redirect(url_for('pcl.admin_tournament_detail', tournament_id=tournament_id))


@pcl.route('/admin/registration/<int:registration_id>/delete', methods=['POST'])
def delete_registration(registration_id):
    """Delete a single player registration"""
    registration = PCLRegistration.query.get_or_404(registration_id)
    team_id = registration.team_id
    player_name = f"{registration.first_name} {registration.last_name}"
    
    try:
        db.session.delete(registration)
        db.session.commit()
        flash(f'Registration "{player_name}" deleted!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting registration: {str(e)}', 'danger')
    
    return redirect(url_for('pcl.admin_team_detail', team_id=team_id))


# ============================================================================
# CAPTAIN MESSAGING (WhatsApp Invitations & Reminders)
# ============================================================================

def get_captain_invitation_message(team, captain_name, captain_url, language='EN'):
    """Get captain invitation message in specified language"""
    messages = {
        'EN': f""" PCL {team.tournament.name} - Team Captain Invitation

Hi {captain_name}! Ã°Å¸â€˜â€¹

You have been selected as Captain for {team.country_flag} {team.country_name} {team.age_category}!

Ã°Å¸â€œâ€¹ Your responsibilities:
Ã¢â‚¬Â¢ Register your team players
Ã¢â‚¬Â¢ Ensure all profiles are complete
Ã¢â‚¬Â¢ Coordinate with your team

Ã°Å¸â€â€” Your secret Captain Dashboard:
{captain_url}

Ã¢Å¡Â Ã¯Â¸Â Keep this link private - only you should have access!

Ã°Å¸â€œâ€¦ Deadline: {team.tournament.registration_deadline.strftime('%d.%m.%Y %H:%M')}

Let's go! 
WPC Series Europe""",
        
        'DE': f""" PCL {team.tournament.name} - Team-Kapitaen Einladung

Hallo {captain_name}! Ã°Å¸â€˜â€¹

Du wurdest als Kapitaen fuer {team.country_flag} {team.country_name} {team.age_category} ausgewaehlt!

Ã°Å¸â€œâ€¹ Deine Aufgaben:
Ã¢â‚¬Â¢ Team-Spieler registrieren
Ã¢â‚¬Â¢ Alle Profile vervollstaendigen
Ã¢â‚¬Â¢ Mit deinem Team koordinieren

Ã°Å¸â€â€” Dein geheimes Kapitaen-Dashboard:
{captain_url}

Ã¢Å¡Â Ã¯Â¸Â Behalte diesen Link privat - nur du solltest Zugriff haben!

Ã°Å¸â€œâ€¦ Deadline: {team.tournament.registration_deadline.strftime('%d.%m.%Y %H:%M')}

Los geht's! 
WPC Series Europe""",
        
        'ES': f""" PCL {team.tournament.name} - InvitaciÃƒÂ³n CapitÃƒÂ¡n

Hola {captain_name}! Ã°Å¸â€˜â€¹

Has sido seleccionado como CapitÃƒÂ¡n de {team.country_flag} {team.country_name} {team.age_category}!

Ã°Å¸â€œâ€¹ Tus responsabilidades:
Ã¢â‚¬Â¢ Registrar los jugadores del equipo
Ã¢â‚¬Â¢ Asegurar que todos los perfiles esten completos
Ã¢â‚¬Â¢ Coordinar con tu equipo

Ã°Å¸â€â€” Tu Panel de CapitÃƒÂ¡n secreto:
{captain_url}

Ã¢Å¡Â Ã¯Â¸Â Manten este enlace privado!

Ã°Å¸â€œâ€¦ Fecha limite: {team.tournament.registration_deadline.strftime('%d.%m.%Y %H:%M')}

Vamos! 
WPC Series Europe""",
        
        'FR': f""" PCL {team.tournament.name} - Invitation Capitaine

Bonjour {captain_name}! Ã°Å¸â€˜â€¹

Vous avez ete selectionne comme Capitaine de {team.country_flag} {team.country_name} {team.age_category}!

Ã°Å¸â€œâ€¹ Vos responsabilites:
Ã¢â‚¬Â¢ Inscrire les joueurs de l'equipe
Ã¢â‚¬Â¢ S'assurer que tous les profils sont complets
Ã¢â‚¬Â¢ Coordonner avec votre equipe

Ã°Å¸â€â€” Votre Tableau de bord Capitaine secret:
{captain_url}

Ã¢Å¡Â Ã¯Â¸Â Gardez ce lien prive!

Ã°Å¸â€œâ€¦ Date limite: {team.tournament.registration_deadline.strftime('%d.%m.%Y %H:%M')}

C'est parti! 
WPC Series Europe"""
    }
    
    return messages.get(language, messages['EN'])


def get_captain_reminder_message(team, captain_name, captain_url, stats, language='EN'):
    """Get captain reminder message in specified language"""
    messages = {
        'EN': f"""Ã¢ÂÂ° PCL {team.tournament.name} - Reminder!

Hi {captain_name}!

Your team {team.country_flag} {team.country_name} {team.age_category} is not yet complete.

Ã°Å¸â€œÅ  Current status:
Ã°Å¸â€˜Â¨ Men: {stats['men']}/{team.min_men}-{team.max_men}
Ã°Å¸â€˜Â© Women: {stats['women']}/{team.min_women}-{team.max_women}
Ã¢Å“â€œ Complete profiles: {stats['men_complete'] + stats['women_complete']}/{stats['total']}

Ã°Å¸â€â€” Captain Dashboard:
{captain_url}

Ã°Å¸â€œâ€¦ Deadline: {team.tournament.registration_deadline.strftime('%d.%m.%Y %H:%M')}

Please complete your team! 
WPC Series Europe""",
        
        'DE': f"""Ã¢ÂÂ° PCL {team.tournament.name} - Erinnerung!

Hallo {captain_name}!

Dein Team {team.country_flag} {team.country_name} {team.age_category} ist noch nicht vollstaendig.

Ã°Å¸â€œÅ  Aktueller Status:
Ã°Å¸â€˜Â¨ Maenner: {stats['men']}/{team.min_men}-{team.max_men}
Ã°Å¸â€˜Â© Frauen: {stats['women']}/{team.min_women}-{team.max_women}
Ã¢Å“â€œ Vollstaendige Profile: {stats['men_complete'] + stats['women_complete']}/{stats['total']}

Ã°Å¸â€â€” Kapitaen-Dashboard:
{captain_url}

Ã°Å¸â€œâ€¦ Deadline: {team.tournament.registration_deadline.strftime('%d.%m.%Y %H:%M')}

Bitte vervollstaendige dein Team! 
WPC Series Europe""",
        
        'ES': f"""Ã¢ÂÂ° PCL {team.tournament.name} - Recordatorio!

Hola {captain_name}!

Tu equipo {team.country_flag} {team.country_name} {team.age_category} aÃƒÂºn no estÃƒÂ¡ completo.

Ã°Å¸â€œÅ  Estado actual:
Ã°Å¸â€˜Â¨ Hombres: {stats['men']}/{team.min_men}-{team.max_men}
Ã°Å¸â€˜Â© Mujeres: {stats['women']}/{team.min_women}-{team.max_women}
Ã¢Å“â€œ Perfiles completos: {stats['men_complete'] + stats['women_complete']}/{stats['total']}

Ã°Å¸â€â€” Panel de CapitÃƒÂ¡n:
{captain_url}

Ã°Å¸â€œâ€¦ Fecha limite: {team.tournament.registration_deadline.strftime('%d.%m.%Y %H:%M')}

Por favor completa tu equipo! 
WPC Series Europe""",
        
        'FR': f"""Ã¢ÂÂ° PCL {team.tournament.name} - Rappel!

Bonjour {captain_name}!

Votre equipe {team.country_flag} {team.country_name} {team.age_category} n'est pas encore complÃƒÂ¨te.

Ã°Å¸â€œÅ  Statut actuel:
Ã°Å¸â€˜Â¨ Hommes: {stats['men']}/{team.min_men}-{team.max_men}
Ã°Å¸â€˜Â© Femmes: {stats['women']}/{team.min_women}-{team.max_women}
Ã¢Å“â€œ Profils complets: {stats['men_complete'] + stats['women_complete']}/{stats['total']}

Ã°Å¸â€â€” Tableau de bord Capitaine:
{captain_url}

Ã°Å¸â€œâ€¦ Date limite: {team.tournament.registration_deadline.strftime('%d.%m.%Y %H:%M')}

Veuillez completer votre equipe! 
WPC Series Europe"""
    }
    
    return messages.get(language, messages['EN'])


@pcl.route('/admin/team/<int:team_id>/send-captain-invite', methods=['GET', 'POST'])
def send_captain_invite(team_id):
    """Send captain invitation via WhatsApp"""
    team = PCLTeam.query.get_or_404(team_id)
    captain_reg = team.registrations.filter_by(is_captain=True).first()
    
    if request.method == 'POST':
        captain_name = request.form.get('captain_name', 'Captain')
        captain_phone = request.form.get('captain_phone', '')
        language = request.form.get('language', 'EN')
        test_mode = request.form.get('test_mode') == 'on'
        
        if not captain_phone:
            flash('Phone number is required!', 'danger')
            return redirect(url_for('pcl.send_captain_invite', team_id=team_id))
        
        captain_url = request.host_url.rstrip('/') + url_for('pcl.captain_dashboard', token=team.captain_token)
        message = get_captain_invitation_message(team, captain_name, captain_url, language)
        sponsor_block = get_whatsapp_sponsor_block(pcl_tournament_id=team.tournament_id, language=language)
        message += sponsor_block

        result = send_whatsapp_message(captain_phone, message, test_mode=test_mode)
        
        if result.get('status') in ['sent', 'queued', 'test_mode']:
            mode_text = " (TEST MODE)" if test_mode else ""
            flash(f'Captain invitation sent to {captain_name}{mode_text}!', 'success')
        else:
            flash(f'Error sending: {result.get("error", "Unknown error")}', 'danger')
        
        return redirect(url_for('pcl.admin_team_detail', team_id=team_id))
    
    return render_template('pcl/send_captain_invite.html', 
                         team=team,
                         captain_reg=captain_reg)


@pcl.route('/admin/team/<int:team_id>/send-captain-reminder', methods=['POST'])
def send_captain_reminder(team_id):
    """Send captain reminder via WhatsApp"""
    team = PCLTeam.query.get_or_404(team_id)
    
    captain_name = request.form.get('captain_name', 'Captain')
    captain_phone = request.form.get('captain_phone', '')
    language = request.form.get('language', 'EN')
    test_mode = request.form.get('test_mode') == 'on'
    
    if not captain_phone:
        flash('Phone number is required!', 'danger')
        return redirect(url_for('pcl.admin_team_detail', team_id=team_id))
    
    captain_url = request.host_url.rstrip('/') + url_for('pcl.captain_dashboard', token=team.captain_token)
    stats = team.get_stats()
    message = get_captain_reminder_message(team, captain_name, captain_url, stats, language)
    sponsor_block = get_whatsapp_sponsor_block(pcl_tournament_id=team.tournament_id, language=language)
    message += sponsor_block

    result = send_whatsapp_message(captain_phone, message, test_mode=test_mode)
    
    if result.get('status') in ['sent', 'queued', 'test_mode']:
        mode_text = " (TEST MODE)" if test_mode else ""
        flash(f'Reminder sent to {captain_name}{mode_text}!', 'success')
    else:
        flash(f'Error sending: {result.get("error", "Unknown error")}', 'danger')
    
    return redirect(url_for('pcl.admin_team_detail', team_id=team_id))


@pcl.route('/admin/tournament/<int:tournament_id>/send-all-reminders', methods=['POST'])
def send_all_captain_reminders(tournament_id):
    """Send reminders to all captains with incomplete teams"""
    tournament = PCLTournament.query.get_or_404(tournament_id)
    test_mode = request.form.get('test_mode') == 'on'
    
    sent_count = 0
    error_count = 0
    skipped_count = 0
    
    for team in tournament.teams.all():
        stats = team.get_stats()
        
        if stats['is_complete']:
            skipped_count += 1
            continue
        
        captain_reg = team.registrations.filter_by(is_captain=True).first()
        
        if not captain_reg or not captain_reg.phone:
            error_count += 1
            continue
        
        captain_url = request.host_url.rstrip('/') + url_for('pcl.captain_dashboard', token=team.captain_token)
        message = get_captain_reminder_message(
            team,
            captain_reg.first_name,
            captain_url,
            stats,
            captain_reg.preferred_language or 'EN'
        )
        sponsor_block = get_whatsapp_sponsor_block(pcl_tournament_id=team.tournament_id, language=captain_reg.preferred_language or 'EN')
        message += sponsor_block

        result = send_whatsapp_message(captain_reg.phone, message, test_mode=test_mode)
        
        if result.get('status') in ['sent', 'queued', 'test_mode']:
            sent_count += 1
        else:
            error_count += 1
    
    mode_text = " (TEST MODE)" if test_mode else ""
    if sent_count > 0:
        flash(f'{sent_count} reminder(s) sent{mode_text}!', 'success')
    if skipped_count > 0:
        flash(f'{skipped_count} complete team(s) skipped.', 'info')
    if error_count > 0:
        flash(f'{error_count} could not be sent (no captain/phone).', 'warning')
    
    return redirect(url_for('pcl.admin_tournament_detail', tournament_id=tournament_id))


# ============================================================================
# API ENDPOINTS
# ============================================================================

@pcl.route('/api/team/<token>/status')
def api_team_status(token):
    """API endpoint for team status (for AJAX updates)"""
    team = PCLTeam.query.filter_by(captain_token=token).first_or_404()
    stats = team.get_stats()
    
    return jsonify({
        'team': f"{team.country_flag} {team.country_name} {team.age_category}",
        'stats': stats,
        'deadline': team.tournament.registration_deadline.isoformat(),
        'is_complete': stats['is_complete']
    })


@pcl.route('/api/team/<token>/players')
def api_team_players(token):
    """API endpoint for team players list"""
    team = PCLTeam.query.filter_by(captain_token=token).first_or_404()
    
    players = []
    for reg in team.registrations.all():
        players.append({
            'id': reg.id,
            'name': f"{reg.first_name} {reg.last_name}",
            'gender': reg.gender,
            'status': reg.status,
            'has_photo': bool(reg.photo_filename),
            'has_bio': bool(reg.bio),
            'shirt_size': reg.shirt_size,
            'shirt_size_2': reg.shirt_size_2,
            'shirt_size_3': reg.shirt_size_3,
            'missing_fields': reg.get_missing_fields()
        })
    
    return jsonify({
        'team': f"{team.country_flag} {team.country_name} {team.age_category}",
        'players': players,
        'stats': team.get_stats()
    })

# ============================================================================
# PLAYER CARDS ROUTES
# ============================================================================
@pcl.route('/cards/tournament/<int:tournament_id>')
def player_cards_tournament(tournament_id):
    """Player cards generator for entire tournament"""
    tournament = PCLTournament.query.get_or_404(tournament_id)
    
    registrations = []
    for team in tournament.teams.order_by(PCLTeam.country_name).all():
        for reg in team.registrations.all():
            registrations.append(reg)
    
    # Use first team as default for template compatibility
    first_team = tournament.teams.first()
    
    return render_template('pcl/player_cards.html',
                         team=first_team,
                         tournament=tournament,
                         registrations=registrations,
                         all_teams=True)

@pcl.route('/cards/team/<int:team_id>')
def player_cards_team(team_id):
    """Player cards generator for a single team"""
    team = PCLTeam.query.get_or_404(team_id)
    registrations = team.registrations.all()
    
    return render_template('pcl/player_cards.html',
                         team=team,
                         tournament=team.tournament,
                         registrations=registrations,
                         all_teams=False)


# ============================================================================
# CHECK-IN SYSTEM
# ============================================================================

# Helper: QR Code Generation
def generate_qr_code_base64(data, size=200):
    """Generate QR code as base64 string"""
    try:
        import qrcode
        from PIL import Image
        
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=4,
        )
        qr.add_data(data)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        img = img.resize((size, size), Image.Resampling.LANCZOS)
        
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        
        return base64.b64encode(buffer.getvalue()).decode('utf-8')
    except ImportError:
        return None


def get_qr_code_url(data, size=200):
    """Get QR code URL (fallback if qrcode library not available)"""
    import urllib.parse
    encoded = urllib.parse.quote(data)
    return f"https://api.qrserver.com/v1/create-qr-code/?size={size}x{size}&data={encoded}"


# Check-in Translations
CHECKIN_TRANSLATIONS = {
    'EN': {
        'checkin_title': 'Tournament Check-in',
        'welcome': 'Welcome',
        'team': 'Team',
        'shirt_size': 'Shirt Size',
        'checkin_button': 'Check In Now',
        'already_checked_in': 'Already Checked In',
        'checked_in_at': 'Checked in at',
        'success_title': 'Check-in Complete!',
        'success_message': 'Please proceed to the Welcome Desk to collect your welcome pack.',
        'show_pass': 'Show this pass at the Welcome Desk',
        'add_to_wallet': 'Add to Apple Wallet',
        'save_pass': 'Save / Screenshot',
        'staff_station': 'Staff Check-in Station',
        'scan_qr': 'Scan QR Code',
        'search_player': 'Search Player',
        'not_found': 'Player not found',
        'checkin_success': 'Successfully checked in',
        'total_players': 'Total Players',
        'checked_in': 'Checked In',
        'pending': 'Pending',
    },
    'DE': {
        'checkin_title': 'Turnier Check-in',
        'welcome': 'Willkommen',
        'team': 'Team',
        'shirt_size': 'Shirt-GrÃ¶ÃŸe',
        'checkin_button': 'Jetzt einchecken',
        'already_checked_in': 'Bereits eingecheckt',
        'checked_in_at': 'Eingecheckt um',
        'success_title': 'Check-in erfolgreich!',
        'success_message': 'Bitte gehe zum Welcome Desk um dein Welcome Pack abzuholen.',
        'show_pass': 'Zeige diesen Pass am Welcome Desk',
        'add_to_wallet': 'Zu Apple Wallet hinzufÃ¼gen',
        'save_pass': 'Speichern / Screenshot',
        'staff_station': 'Staff Check-in Station',
        'scan_qr': 'QR-Code scannen',
        'search_player': 'Spieler suchen',
        'not_found': 'Spieler nicht gefunden',
        'checkin_success': 'Erfolgreich eingecheckt',
        'total_players': 'Gesamt Spieler',
        'checked_in': 'Eingecheckt',
        'pending': 'Ausstehend',
    },
    'ES': {
        'checkin_title': 'Check-in del Torneo',
        'welcome': 'Bienvenido',
        'team': 'Equipo',
        'shirt_size': 'Talla de Camiseta',
        'checkin_button': 'Hacer Check-in',
        'already_checked_in': 'Ya registrado',
        'checked_in_at': 'Registrado a las',
        'success_title': 'Â¡Check-in completado!',
        'success_message': 'Por favor dirÃ­gete al Welcome Desk para recoger tu pack de bienvenida.',
        'show_pass': 'Muestra este pase en el Welcome Desk',
        'add_to_wallet': 'AÃ±adir a Apple Wallet',
        'save_pass': 'Guardar / Captura',
        'staff_station': 'EstaciÃ³n de Check-in Staff',
        'scan_qr': 'Escanear cÃ³digo QR',
        'search_player': 'Buscar jugador',
        'not_found': 'Jugador no encontrado',
        'checkin_success': 'Check-in exitoso',
        'total_players': 'Total Jugadores',
        'checked_in': 'Registrados',
        'pending': 'Pendientes',
    },
    'FR': {
        'checkin_title': 'Check-in du Tournoi',
        'welcome': 'Bienvenue',
        'team': 'Ã‰quipe',
        'shirt_size': 'Taille de Maillot',
        'checkin_button': 'Check-in maintenant',
        'already_checked_in': 'DÃ©jÃ  enregistrÃ©',
        'checked_in_at': 'EnregistrÃ© Ã ',
        'success_title': 'Check-in rÃ©ussi!',
        'success_message': 'Veuillez vous rendre au Welcome Desk pour rÃ©cupÃ©rer votre pack de bienvenue.',
        'show_pass': 'Montrez ce pass au Welcome Desk',
        'add_to_wallet': 'Ajouter Ã  Apple Wallet',
        'save_pass': 'Sauvegarder / Capture',
        'staff_station': 'Station de Check-in Staff',
        'scan_qr': 'Scanner le code QR',
        'search_player': 'Rechercher un joueur',
        'not_found': 'Joueur non trouvÃ©',
        'checkin_success': 'Check-in rÃ©ussi',
        'total_players': 'Total Joueurs',
        'checked_in': 'EnregistrÃ©s',
        'pending': 'En attente',
    }
}

def get_checkin_translations(lang='EN'):
    return CHECKIN_TRANSLATIONS.get(lang, CHECKIN_TRANSLATIONS['EN'])


# ============================================================================
# PLAYER SELF CHECK-IN ROUTES
# ============================================================================

@pcl.route('/checkin/<token>')
def player_checkin(token):
    """Player check-in page - accessed via QR code or link"""
    registration = PCLRegistration.query.filter_by(profile_token=token).first_or_404()
    team = registration.team
    tournament = team.tournament
    
    lang = request.args.get('lang', registration.preferred_language or 'EN').upper()
    if lang not in CHECKIN_TRANSLATIONS:
        lang = 'EN'
    t = get_checkin_translations(lang)
    
    # Generate QR code
    checkin_url = request.host_url.rstrip('/') + url_for('pcl.player_checkin', token=token)
    qr_base64 = generate_qr_code_base64(checkin_url, size=250)
    qr_url = get_qr_code_url(checkin_url, size=250) if not qr_base64 else None
    
    return render_template('pcl/checkin.html',
                         registration=registration,
                         team=team,
                         tournament=tournament,
                         qr_base64=qr_base64,
                         qr_url=qr_url,
                         t=t,
                         current_lang=lang)


@pcl.route('/checkin/<token>/confirm', methods=['POST'])
def confirm_checkin(token):
    """Process player check-in with phone and GDPR consent"""
    registration = PCLRegistration.query.filter_by(profile_token=token).first_or_404()

    lang = request.form.get('language', registration.preferred_language or 'EN').upper()

    # Check if already checked in
    if registration.checked_in:
        return redirect(url_for('pcl.wallet_pass', token=token, lang=lang))

    # Get form data
    phone = request.form.get('phone', '').strip()
    privacy_accepted = request.form.get('privacy_accepted') == 'on'
    whatsapp_optin = request.form.get('whatsapp_optin') == 'on'
    marketing_optin = request.form.get('marketing_optin') == 'on'

    # Validate required fields
    if not phone:
        flash('Phone number is required', 'danger')
        return redirect(url_for('pcl.player_checkin', token=token))

    if not privacy_accepted:
        flash('You must accept the Privacy Policy', 'danger')
        return redirect(url_for('pcl.player_checkin', token=token))

    # Save consent and phone
    registration.phone = phone
    registration.preferred_language = lang
    registration.privacy_accepted = True
    registration.privacy_accepted_at = datetime.utcnow()
    registration.whatsapp_optin = whatsapp_optin
    registration.marketing_optin = marketing_optin

    # Perform check-in
    registration.checked_in = True
    registration.checked_in_at = datetime.utcnow()

    try:
        db.session.commit()
        return redirect(url_for('pcl.wallet_pass', token=token, lang=lang))
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('pcl.player_checkin', token=token))
# ============================================================================
# WALLET PASS (Web Version)
# ============================================================================

@pcl.route('/checkin/pass/<token>')
def wallet_pass(token):
    """Show wallet-style pass after check-in"""
    registration = PCLRegistration.query.filter_by(profile_token=token).first_or_404()
    team = registration.team
    tournament = team.tournament
    
    lang = request.args.get('lang', registration.preferred_language or 'EN').upper()
    if lang not in CHECKIN_TRANSLATIONS:
        lang = 'EN'
    t = get_checkin_translations(lang)
    
    # Generate QR code
    checkin_url = request.host_url.rstrip('/') + url_for('pcl.player_checkin', token=token)
    qr_base64 = generate_qr_code_base64(checkin_url, size=180)
    qr_url = get_qr_code_url(checkin_url, size=180) if not qr_base64 else None
    
    return render_template('pcl/wallet_pass.html',
                         registration=registration,
                         team=team,
                         tournament=tournament,
                         qr_base64=qr_base64,
                         qr_url=qr_url,
                         t=t,
                         current_lang=lang)


# ============================================================================
# APPLE WALLET (.pkpass)
# ============================================================================

@pcl.route('/checkin/pass/<token>/apple')
def apple_wallet_pass(token):
    """Generate Apple Wallet .pkpass file"""
    registration = PCLRegistration.query.filter_by(profile_token=token).first_or_404()
    team = registration.team
    tournament = team.tournament
    
    try:
        from utils.wallet_pass import create_pkpass, is_apple_wallet_available
        
        if not is_apple_wallet_available():
            flash('Apple Wallet is not configured', 'warning')
            return redirect(url_for('pcl.wallet_pass', token=token))
        
        pkpass_data = create_pkpass(
            registration=registration,
            team=team,
            tournament=tournament
        )
        
        if pkpass_data:
            filename = f"WPC_{registration.first_name}_{registration.last_name}.pkpass"
            return send_file(
                io.BytesIO(pkpass_data),
                mimetype='application/vnd.apple.pkpass',
                as_attachment=True,
                download_name=filename
            )
        else:
            flash('Error generating Apple Wallet pass', 'danger')
            return redirect(url_for('pcl.wallet_pass', token=token))
            
    except ImportError:
        flash('Apple Wallet module not available', 'warning')
        return redirect(url_for('pcl.wallet_pass', token=token))
    except Exception as e:
        print(f"Apple Wallet Error: {str(e)}")
        flash(f'Error: {str(e)}', 'danger')
        return redirect(url_for('pcl.wallet_pass', token=token))


# ============================================================================
# ADMIN: STAFF CHECK-IN STATION
# ============================================================================

@pcl.route('/admin/tournament/<int:tournament_id>/checkin')
def staff_checkin_station(tournament_id):
    """Staff check-in station dashboard"""
    tournament = PCLTournament.query.get_or_404(tournament_id)
    
    # Get all registrations for this tournament
    registrations = PCLRegistration.query.join(PCLTeam).filter(
        PCLTeam.tournament_id == tournament_id
    ).order_by(PCLRegistration.last_name).all()
    
    # Stats
    total = len(registrations)
    checked_in = len([r for r in registrations if r.checked_in])
    pending = total - checked_in
    
    # Group by team
    teams_data = {}
    for reg in registrations:
        team_key = f"{reg.team.country_flag} {reg.team.country_name} {reg.team.age_category}"
        if team_key not in teams_data:
            teams_data[team_key] = {
                'team': reg.team,
                'players': [],
                'checked_in': 0,
                'total': 0
            }
        teams_data[team_key]['players'].append(reg)
        teams_data[team_key]['total'] += 1
        if reg.checked_in:
            teams_data[team_key]['checked_in'] += 1
    
    t = get_checkin_translations('EN')
    
    return render_template('pcl/staff_checkin.html',
                         tournament=tournament,
                         registrations=registrations,
                         teams_data=teams_data,
                         stats={'total': total, 'checked_in': checked_in, 'pending': pending},
                         t=t)


@pcl.route('/admin/tournament/<int:tournament_id>/checkin/search')
def staff_search_player(tournament_id):
    """Search for player (AJAX endpoint)"""
    query = request.args.get('q', '').strip().lower()
    
    if len(query) < 2:
        return jsonify({'players': []})
    
    registrations = PCLRegistration.query.join(PCLTeam).filter(
        PCLTeam.tournament_id == tournament_id,
        db.or_(
            PCLRegistration.first_name.ilike(f'%{query}%'),
            PCLRegistration.last_name.ilike(f'%{query}%')
        )
    ).limit(10).all()

    players = []
    for reg in registrations:
        players.append({
            'id': reg.id,
            'name': f"{reg.first_name} {reg.last_name}",
            'team': f"{reg.team.country_flag} {reg.team.country_name} {reg.team.age_category}",
            'shirt_size': reg.shirt_size,
            'shirt_size_2': reg.shirt_size_2,
            'shirt_size_3': reg.shirt_size_3,
            'checked_in': reg.checked_in,
            'checked_in_at': reg.checked_in_at.strftime('%H:%M') if reg.checked_in_at else None,
            'photo': reg.photo_filename
        })
    
    return jsonify({'players': players})


@pcl.route('/admin/checkin/<int:registration_id>', methods=['POST'])
def staff_do_checkin(registration_id):
    """Staff performs check-in for a player"""
    registration = PCLRegistration.query.get_or_404(registration_id)
    
    if not registration.checked_in:
        registration.checked_in = True
        registration.checked_in_at = datetime.utcnow()
        
        try:
            db.session.commit()
            return jsonify({
                'success': True,
                'message': f'{registration.first_name} {registration.last_name} checked in!',
                'checked_in_at': registration.checked_in_at.strftime('%H:%M')
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    else:
        return jsonify({
            'success': True,
            'message': 'Already checked in',
            'checked_in_at': registration.checked_in_at.strftime('%H:%M') if registration.checked_in_at else None
        })


@pcl.route('/admin/checkin/<int:registration_id>/undo', methods=['POST'])
def staff_undo_checkin(registration_id):
    """Undo a check-in (staff only)"""
    registration = PCLRegistration.query.get_or_404(registration_id)
    
    registration.checked_in = False
    registration.checked_in_at = None
    
    try:
        db.session.commit()
        return jsonify({'success': True, 'message': 'Check-in undone'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# ADMIN: QR CODE GENERATION
# ============================================================================

@pcl.route('/admin/tournament/<int:tournament_id>/qrcodes')
def generate_qrcodes(tournament_id):
    """Generate printable QR codes for all players"""
    tournament = PCLTournament.query.get_or_404(tournament_id)
    
    registrations = PCLRegistration.query.join(PCLTeam).filter(
        PCLTeam.tournament_id == tournament_id
    ).order_by(PCLTeam.country_name, PCLRegistration.last_name).all()
    
    # Generate tokens if missing
    for reg in registrations:
        if not reg.profile_token:
            reg.generate_profile_token()
    
    try:
        db.session.commit()
    except:
        db.session.rollback()
    
    # Generate QR codes
    players_with_qr = []
    base_url = request.host_url.rstrip('/')
    
    for reg in registrations:
        checkin_url = f"{base_url}/pcl/checkin/{reg.profile_token}"
        qr_base64 = generate_qr_code_base64(checkin_url, size=150)
        qr_url = get_qr_code_url(checkin_url, size=150) if not qr_base64 else None
        
        players_with_qr.append({
            'registration': reg,
            'qr_base64': qr_base64,
            'qr_url': qr_url,
            'checkin_url': checkin_url
        })
    
    return render_template('pcl/qrcodes_print.html',
                         tournament=tournament,
                         players=players_with_qr)


# ============================================================================
# API: LIVE STATS
# ============================================================================

@pcl.route('/api/checkin/stats/<int:tournament_id>')
def checkin_stats_api(tournament_id):
    """Get live check-in stats (for dashboard auto-refresh)"""
    registrations = PCLRegistration.query.join(PCLTeam).filter(
        PCLTeam.tournament_id == tournament_id
    ).all()
    
    total = len(registrations)
    checked_in = len([r for r in registrations if r.checked_in])
    
    # Recent check-ins
    recent = PCLRegistration.query.join(PCLTeam).filter(
        PCLTeam.tournament_id == tournament_id,
        PCLRegistration.checked_in == True
    ).order_by(PCLRegistration.checked_in_at.desc()).limit(5).all()
    
    return jsonify({
        'total': total,
        'checked_in': checked_in,
        'pending': total - checked_in,
        'percentage': round(checked_in / total * 100, 1) if total > 0 else 0,
        'recent': [
            {
                'name': f"{r.first_name} {r.last_name}",
                'team': f"{r.team.country_flag} {r.team.country_name}",
                'time': r.checked_in_at.strftime('%H:%M') if r.checked_in_at else None
            }
            for r in recent
        ]
    })
@pcl.route('/terms')
def terms():
    """Privacy Policy and Terms page"""
    return render_template('terms.html')


# ============================================================================
# PCL MATCH / LINEUP MODULE - PHASE 2 (Lineup Submission: Captain + Admin)
# ============================================================================

LINEUP_FIELDS = [
    'wd_player1_id', 'wd_player2_id', 'md_player1_id', 'md_player2_id',
    'mx1_male_id', 'mx1_female_id', 'mx2_male_id', 'mx2_female_id',
    'hb_player1_id', 'hb_player2_id', 'hb_player3_id', 'hb_player4_id',
]


def _selectable_players(team, extra_ids=None):
    """Selectable lineup players for a team: every playing member (non-playing
    captains excluded), plus any extra ids already stored in a lineup so removed
    players still render. Returns (all, males, females), each name-sorted."""
    regs = team.registrations.all()
    selectable = [r for r in regs if not (r.is_captain and not r.is_playing)]
    have = {r.id for r in selectable}
    for rid in (extra_ids or []):
        if rid and rid not in have:
            r = PCLRegistration.query.get(rid)
            if r:
                selectable.append(r)
                have.add(rid)
    keyfn = lambda r: ((r.first_name or '').lower(), (r.last_name or '').lower())
    selectable.sort(key=keyfn)
    males = sorted([r for r in selectable if r.gender == 'male'], key=keyfn)
    females = sorted([r for r in selectable if r.gender == 'female'], key=keyfn)
    return selectable, males, females


def _lineup_page(match, team, is_admin_mode, lang, t, token=None, submitted_by=None):
    """Shared GET/POST handler for the lineup form (captain and admin)."""
    existing = match.lineups.filter_by(team_id=team.id).first()

    if request.method == 'POST':
        ids = {f: request.form.get(f, type=int) for f in LINEUP_FIELDS}
        lineup = existing or PCLLineup(match_id=match.id, team_id=team.id)
        lineup.team_id = team.id
        for f in LINEUP_FIELDS:
            setattr(lineup, f, ids[f])

        with db.session.no_autoflush:
            errors = lineup.validate()

        if errors:
            db.session.rollback()
            for e in errors:
                flash(e, 'danger')
            selected = ids
        else:
            now = datetime.now()
            if not lineup.submitted_at:
                lineup.submitted_at = now
            lineup.last_modified_at = now
            # Captain submission stores the captain id; admin override leaves it
            # NULL (with submitted_at set) to flag an admin-entered lineup.
            lineup.submitted_by_captain_id = submitted_by
            if existing is None:
                db.session.add(lineup)
            try:
                db.session.commit()
                flash(t['lineup_already_submitted'], 'success')
                if is_admin_mode:
                    return redirect(url_for('pcl.admin_match_detail', match_id=match.id, lang=lang))
                return redirect(url_for('pcl.captain_dashboard', token=token, lang=lang))
            except Exception as e:
                db.session.rollback()
                flash(f'Error: {str(e)}', 'danger')
                selected = ids
    else:
        if existing:
            selected = {f: getattr(existing, f) for f in LINEUP_FIELDS}
        else:
            selected = {f: None for f in LINEUP_FIELDS}

    extra = [v for v in selected.values() if v]
    all_players, males, females = _selectable_players(team, extra_ids=extra)

    is_home = match.team_home_id == team.id
    opp_team = match.team_away if is_home else match.team_home
    opp_id = match.team_away_id if is_home else match.team_home_id
    opp_lineup = match.lineups.filter_by(team_id=opp_id).first()

    if is_admin_mode:
        back_url = url_for('pcl.admin_match_detail', match_id=match.id, lang=lang)
    else:
        back_url = url_for('pcl.captain_dashboard', token=token, lang=lang)

    return render_template('pcl/lineup_submit.html',
                         match=match, team=team, opp_team=opp_team,
                         is_admin_mode=is_admin_mode,
                         existing=existing, selected=selected,
                         players=all_players, males=males, females=females,
                         opp_submitted=bool(opp_lineup and opp_lineup.is_submitted()),
                         deadline_passed=match.is_lineup_deadline_passed(),
                         back_url=back_url, token=token,
                         t=t, current_lang=lang)


@pcl.route('/team/<token>/match/<int:match_id>/lineup', methods=['GET', 'POST'])
def captain_lineup(token, match_id):
    """Captain: submit/edit their team's lineup (editable until the deadline)."""
    team = PCLTeam.query.filter_by(captain_token=token).first_or_404()
    match = PCLMatch.query.get_or_404(match_id)
    lang = _match_lang()
    t = get_translations(lang)

    if team.id not in (match.team_home_id, match.team_away_id):
        flash('This match does not involve your team.', 'danger')
        return redirect(url_for('pcl.captain_dashboard', token=token, lang=lang))

    if match.is_lineup_deadline_passed():
        flash(t['lineup_deadline_passed'], 'warning')
        return redirect(url_for('pcl.captain_lineup_reveal', token=token, match_id=match_id, lang=lang))

    captain = team.registrations.filter_by(is_captain=True).first()
    return _lineup_page(match, team, is_admin_mode=False, lang=lang, t=t,
                        token=token, submitted_by=(captain.id if captain else None))


@pcl.route('/team/<token>/match/<int:match_id>/reveal')
def captain_lineup_reveal(token, match_id):
    """Captain: view both lineups (only after the deadline has passed)."""
    team = PCLTeam.query.filter_by(captain_token=token).first_or_404()
    match = PCLMatch.query.get_or_404(match_id)
    lang = _match_lang()
    t = get_translations(lang)

    if team.id not in (match.team_home_id, match.team_away_id):
        flash('This match does not involve your team.', 'danger')
        return redirect(url_for('pcl.captain_dashboard', token=token, lang=lang))

    if not match.is_lineup_deadline_passed():
        flash('Lineups are revealed after the deadline.', 'info')
        return redirect(url_for('pcl.captain_dashboard', token=token, lang=lang))

    home_lineup = match.lineups.filter_by(team_id=match.team_home_id).first()
    away_lineup = match.lineups.filter_by(team_id=match.team_away_id).first()
    return render_template('pcl/lineup_reveal.html',
                         match=match, home_lineup=home_lineup, away_lineup=away_lineup,
                         is_admin_mode=False,
                         back_url=url_for('pcl.captain_dashboard', token=token, lang=lang),
                         t=t, current_lang=lang)


@pcl.route('/admin/match/<int:match_id>/lineup/<int:team_id>', methods=['GET', 'POST'])
def admin_match_lineup(match_id, team_id):
    """Admin: submit/edit a lineup for either team, before or after the deadline."""
    match = PCLMatch.query.get_or_404(match_id)
    team = PCLTeam.query.get_or_404(team_id)
    lang = _match_lang()
    t = get_translations(lang)

    if team.id not in (match.team_home_id, match.team_away_id):
        flash('That team is not part of this match.', 'danger')
        return redirect(url_for('pcl.admin_match_detail', match_id=match.id, lang=lang))

    return _lineup_page(match, team, is_admin_mode=True, lang=lang, t=t,
                        token=None, submitted_by=None)


@pcl.route('/admin/match/<int:match_id>/reveal')
def admin_match_reveal(match_id):
    """Admin: view both lineups at any time."""
    match = PCLMatch.query.get_or_404(match_id)
    lang = _match_lang()
    t = get_translations(lang)
    home_lineup = match.lineups.filter_by(team_id=match.team_home_id).first()
    away_lineup = match.lineups.filter_by(team_id=match.team_away_id).first()
    return render_template('pcl/lineup_reveal.html',
                         match=match, home_lineup=home_lineup, away_lineup=away_lineup,
                         is_admin_mode=True,
                         back_url=url_for('pcl.admin_match_detail', match_id=match.id, lang=lang),
                         t=t, current_lang=lang)


# ============================================================================
# PCL MATCH / LINEUP MODULE - PHASE 3 (Score Entry + auto Heartbreaker)
# ============================================================================

SCORE_TYPES = ['wd', 'md', 'mx1', 'mx2', 'hb']


def _upsert_result(match, mtype, home, away):
    """Insert/update one PCLMatchResult row and set its winner ('home'/'away'/None)."""
    r = PCLMatchResult.query.filter_by(match_id=match.id, match_type=mtype).first()
    if r is None:
        r = PCLMatchResult(match_id=match.id, match_type=mtype)
        db.session.add(r)
    r.home_score = home
    r.away_score = away
    if home is not None and away is not None and home > away:
        r.winner = 'home'
    elif home is not None and away is not None and away > home:
        r.winner = 'away'
    else:
        r.winner = None
    r.recorded_at = datetime.now()
    return r


def _recalculate_match(match):
    """Recompute match standing, winner and status from the stored results."""
    final_h, final_a = match.get_final_standing()
    match.home_score = final_h
    match.away_score = final_a
    winner = match.get_winner()
    if winner == 'home':
        match.winner_id = match.team_home_id
        match.status = 'completed'
    elif winner == 'away':
        match.winner_id = match.team_away_id
        match.status = 'completed'
    else:
        match.winner_id = None
        # Any decided result (or a pending heartbreaker) means the match is live.
        if match.get_doubles_count_scored() > 0 or match.is_heartbreaker_required():
            match.status = 'in_progress'
        elif match.status not in ('pending', 'lineups_locked'):
            match.status = 'in_progress'


def _score_page(match, is_admin_mode, lang, t, token=None):
    """Shared GET/POST handler for the score-entry form (captain + admin)."""
    if request.method == 'POST':
        for ty in SCORE_TYPES:
            home_raw = request.form.get(ty + '_home')
            away_raw = request.form.get(ty + '_away')
            # Only touch a row when at least one side was provided for it.
            if (home_raw is None or home_raw == '') and (away_raw is None or away_raw == ''):
                continue
            home = request.form.get(ty + '_home', type=int)
            away = request.form.get(ty + '_away', type=int)
            _upsert_result(match, ty, home, away)

        _recalculate_match(match)
        try:
            db.session.commit()
            flash(t['score_save'], 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')

        if is_admin_mode:
            return redirect(url_for('pcl.admin_match_score', match_id=match.id, lang=lang))
        return redirect(url_for('pcl.captain_match_score', token=token, match_id=match.id, lang=lang))

    # GET
    res = match.results_by_type()
    scores = {}
    for ty in SCORE_TYPES:
        r = res.get(ty)
        scores[ty] = {
            'home': r.home_score if r and r.home_score is not None else '',
            'away': r.away_score if r and r.away_score is not None else '',
            'winner': r.winner if r else None,
        }

    home_lineup = match.lineups.filter_by(team_id=match.team_home_id).first()
    away_lineup = match.lineups.filter_by(team_id=match.team_away_id).first()
    standing = match.get_final_standing()
    winner = match.get_winner()

    if is_admin_mode:
        back_url = url_for('pcl.admin_match_detail', match_id=match.id, lang=lang)
    else:
        back_url = url_for('pcl.captain_dashboard', token=token, lang=lang)

    return render_template('pcl/lineup_score.html',
                         match=match, is_admin_mode=is_admin_mode,
                         home_lineup=home_lineup, away_lineup=away_lineup,
                         scores=scores, standing=standing,
                         doubles_scored=match.get_doubles_count_scored(),
                         hb_required=match.is_heartbreaker_required(),
                         winner=winner, back_url=back_url, token=token,
                         t=t, current_lang=lang)


@pcl.route('/admin/match/<int:match_id>/score', methods=['GET', 'POST'])
def admin_match_score(match_id):
    """Admin: enter scores for a match (anytime)."""
    match = PCLMatch.query.get_or_404(match_id)
    lang = _match_lang()
    t = get_translations(lang)
    return _score_page(match, is_admin_mode=True, lang=lang, t=t, token=None)


@pcl.route('/team/<token>/match/<int:match_id>/score', methods=['GET', 'POST'])
def captain_match_score(token, match_id):
    """Captain: enter scores for their match (only after the lineup deadline)."""
    team = PCLTeam.query.filter_by(captain_token=token).first_or_404()
    match = PCLMatch.query.get_or_404(match_id)
    lang = _match_lang()
    t = get_translations(lang)

    if team.id not in (match.team_home_id, match.team_away_id):
        flash('This match does not involve your team.', 'danger')
        return redirect(url_for('pcl.captain_dashboard', token=token, lang=lang))

    if not match.is_lineup_deadline_passed():
        flash('Scores can be entered after the lineup deadline.', 'info')
        return redirect(url_for('pcl.captain_dashboard', token=token, lang=lang))

    return _score_page(match, is_admin_mode=False, lang=lang, t=t, token=token)


# ============================================================================
# PCL MATCH / LINEUP MODULE - PHASE 4 (Admin WhatsApp message generator)
# ============================================================================

def _team_captain_name(team):
    """Name of the team's captain, falling back to a generic label."""
    cap = team.registrations.filter_by(is_captain=True).first()
    if cap:
        name = f"{cap.first_name or ''} {cap.last_name or ''}".strip()
        if name:
            return name
    return f"{team.country_name} captain"


def _match_whatsapp_data(match):
    """Build the data dict for one match used by the WhatsApp text templates.

    Captain links point straight at this match's lineup submission form (saving a
    tap) and use the current host so they work in any environment. After the
    deadline the captain_lineup route redirects to the reveal page automatically.
    """
    base = request.host_url.rstrip('/')
    home, away = match.team_home, match.team_away
    home_team = f"{home.country_name} {home.age_category}"
    away_team = f"{away.country_name} {away.age_category}"
    return {
        'id': match.id,
        'home_flag': home.country_flag or '',
        'away_flag': away.country_flag or '',
        'home_team': home_team,
        'away_team': away_team,
        'home_captain': _team_captain_name(home),
        'away_captain': _team_captain_name(away),
        'label': f"{home.country_flag} {home_team} vs {away.country_flag} {away_team}",
        # Match line uses weekday + date; deadline line is short (no weekday).
        'date': format_match_date(match.match_date, 'EN') if match.match_date else '',
        'deadline': match.lineup_deadline.strftime('%d.%m. %H:%M') if match.lineup_deadline else '',
        'court': match.court or '',
        'home_url': base + url_for('pcl.captain_lineup', token=home.captain_token, match_id=match.id),
        'away_url': base + url_for('pcl.captain_lineup', token=away.captain_token, match_id=match.id),
    }


@pcl.route('/admin/tournament/<int:tournament_id>/whatsapp')
def admin_whatsapp_tools(tournament_id):
    """Admin: WhatsApp message generator for all matches in a tournament."""
    tournament = PCLTournament.query.get_or_404(tournament_id)
    matches = PCLMatch.query.filter_by(tournament_id=tournament_id).all()
    matches.sort(key=lambda m: (m.match_date is None, m.match_date or datetime.max))
    match_data = [_match_whatsapp_data(m) for m in matches]
    return render_template('pcl/admin_whatsapp_tools.html',
                         tournament=tournament,
                         match_data=match_data,
                         preselect_id=None,
                         back_url=url_for('pcl.admin_match_list', tournament_id=tournament.id))


@pcl.route('/admin/match/<int:match_id>/whatsapp')
def admin_match_whatsapp(match_id):
    """Admin: WhatsApp generator focused on a single match (same page, preselected)."""
    match = PCLMatch.query.get_or_404(match_id)
    tournament = match.tournament
    matches = PCLMatch.query.filter_by(tournament_id=tournament.id).all()
    matches.sort(key=lambda m: (m.match_date is None, m.match_date or datetime.max))
    match_data = [_match_whatsapp_data(m) for m in matches]
    return render_template('pcl/admin_whatsapp_tools.html',
                         tournament=tournament,
                         match_data=match_data,
                         preselect_id=match.id,
                         back_url=url_for('pcl.admin_match_detail', match_id=match.id))

